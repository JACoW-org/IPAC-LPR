import jacow_nd_func as jnf
import openpyxl
import datetime
import matplotlib.pyplot as plt

date_reference=datetime.datetime(2023,5,8,00,59)
days_reviewers_assigned=[]
days_response=[]
interval_response=[]
days_review=[]
interval_assign_review=[]
interval_response_review=[]

if 1==1:
    file = jnf.papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()
    
    irow=2
    while not referees_assignation_wb.cell(row=irow,column=1).value is None:
        print("Checking row", irow)
        date_assigned=datetime.datetime.strptime(referees_assignation_wb.cell(row=irow,column=5).value,"%Y-%m-%d %H:%M:%S")
        days_reviewers_assigned.append((date_assigned-date_reference).days)
        date_response_txt=referees_assignation_wb.cell(row=irow,column=9).value
        if date_response_txt is not None:
            date_response=datetime.datetime.strptime(date_response_txt,"%Y-%m-%d %H:%M:%S")
            days_response.append((date_response-date_reference).days)
            if (date_response-date_assigned).days>0:
                interval_response.append((date_response-date_assigned).days-1)
            #if (date_response-date_assigned).days <0:
            #print("(date_response-date_assigned).days",(date_response-date_assigned).days)
        date_review_txt=referees_assignation_wb.cell(row=irow,column=12).value
        if date_review_txt is not None and len(date_review_txt)>5 and referees_assignation_wb.cell(row=irow,column=6).value=="Review received":
            date_review=datetime.datetime.strptime(referees_assignation_wb.cell(row=irow,column=12).value.split(".")[0],"%Y-%m-%dT%H:%M:%S")
            days_review.append((date_review-date_reference).days)
            if (date_review-date_assigned).days>0:
                interval_assign_review.append((date_review-date_assigned).days-1)
                if date_response_txt is not None and len(date_response_txt)>5:
                    if (date_review-date_response).days >0:
                        interval_response_review.append((date_review-date_response).days-1)
        irow=irow+1

bins_date=range(-60,5,1)
bins_date=range(-60,5,7)
bins_interval=range(-10,30,1)

if 1==1:
    plt.figure(figsize=(4, 3), dpi=100)    
    plt.hist(days_reviewers_assigned,bins=bins_date)
    plt.title('Date assigned')
    plt.xlabel('Days before the conference')
    plt.ylabel('N assignations/day')
    plt.savefig("assignations_per_day.png")
    plt.grid()
    plt.show(block=False)

if 1==1:
    plt.figure(figsize=(4, 3), dpi=100)    
    plt.hist(days_response,bins=bins_date)
    plt.title('Date response')
    plt.xlabel('Days before the conference')
    plt.ylabel('N responses/day')
    plt.savefig("responses_per_day.png")
    plt.grid()
    plt.show(block=False)


if 1==1:
    plt.figure(figsize=(4, 3), dpi=100)    
    plt.hist(interval_response,bins=bins_interval)
    plt.title('Time to respond')
    plt.xlabel('Days between assignation and response')
    plt.ylabel('N responses in the interval')
    plt.savefig("responses_per_interval.png")
    plt.grid()
    plt.show(block=False)

if 1==1:
    plt.figure(figsize=(4, 3), dpi=100)    
    plt.hist(days_review,bins=bins_date)
    plt.title('Date review')
    plt.xlabel('Days before the conference')
    plt.ylabel('N reviews/day')
    plt.savefig("reviews_per_day.png")
    plt.grid()
    plt.show(block=False)


if 1==1:
    plt.figure(figsize=(4, 3), dpi=100)    
    plt.hist(interval_assign_review,bins=bins_interval)
    plt.title('Time to review')
    plt.xlabel('Days between assignation and review')
    plt.ylabel('N reviews in the interval')
    plt.savefig("reviews_per_interval.png")
    plt.grid()
    plt.show(block=False)

if 1==1:
    plt.figure(figsize=(4, 3), dpi=100)    
    plt.hist(interval_response_review,bins=bins_interval)
    plt.title('Time to review after acceptance')
    plt.xlabel('Days between acceptance and review')
    plt.ylabel('N reviews in the interval')
    plt.savefig("review_accepted_per_interval.png")
    plt.grid()
    plt.show(block=False)
plt.show(block=True)

#assignations per paper
#declined per paper
