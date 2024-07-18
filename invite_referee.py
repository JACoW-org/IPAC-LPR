#!/usr/local/bin/python3

# Invites a participant to become a referee
# Nicolas Delerue, 4/2023

import jacow_nd_func as jnf
import email_func as ef 
import argparse
import openpyxl
import datetime
#import urllib.parse
#import requests

parser = argparse.ArgumentParser()
parser.add_argument("--paper", nargs=1, help="The paper to be assigned")
parser.add_argument("--email", nargs=1, help="The email of the participant to be invited.")
parser.add_argument("--override", action="store_true", help="Override limits on referees per paper")

parser.parse_args()
args = parser.parse_args()

if args.paper is not None:
    paper_id = args.paper[0]
    contrib=jnf.find_contrib(the_id=paper_id)
    paper_db_id=contrib['db_id']
else:
    print("No paper specified")
    exit()

[n_referees,same_paper]=jnf.check_referees_for_paper(args.paper[0])
print('n_referees',n_referees,'same_paper',same_paper)

if n_referees>0:
    print("The same paper already has the following referees",n_referees)
    print(same_paper)
    if n_referees>=2:
        print("There are already ",n_referees," referees on this paper")
        if args.override:
            print("Overriding paper limitation")
        else:
            exit()
        
### Trying to invite an additional referee
#Check that the name or the email are not of a known referee

def check_and_add_participant_as_new_referee(email):
    the_ref=jnf.get_referee_by_email(email)
    if the_ref is None:
        ref_id=jnf.find_referee_by_email(email)
        if ref_id is not None and len(ref_id)>0:
            the_ref=jnf.get_referee_by_id(ref_id[0])
        else:
            the_ref=None

    if the_ref is not None:
        print("Referee exists",the_ref)
        return None
    else:
        print("Referee does no exists, looking for participant")
        part=jnf.find_participant_by_email(email)
        if part is None:
            print("Participant not found")
            exit()
            return None
        else:
            print("Participant found",part)
            participant=part
            user=jnf.search_user(email=participant['email'])
            #print(user)
            #Check that the referee is in the list
            #print(jnf.get_referees_list())
            if not str(user['id']) in jnf.get_referees_list():
                print("You must first add ", user['email']," ",user['id'] ," in the team of content_reviewers at https://indico.jacow.org/event/41/manage/papers/")
                exit()
            else:
                print('User ',user['id']," ",user['full_name']," is in the content reviewers list")
            participant['db_id']=user['id']
            participant=add_participant_as_referee(participant)
            return participant


def add_participant_as_referee(participant):
    if jnf.list_referees_wb is None:
        jnf.load_referee_files()
    
    col_for_id=jnf.col_for_id
    col_for_name=jnf.col_for_name
    col_for_country=jnf.col_for_country
    col_for_categ=jnf.col_for_categ
    col_for_MC=jnf.col_for_MC
    col_for_email=jnf.col_for_email
    col_for_affiliation=jnf.col_for_affiliation
    
    file = "all_referees.xlsx"
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    list_referees_wb = wb_obj.active
    irow=2
    while (list_referees_wb.cell(row=irow,column=1).value is not None):
        irow=irow+1
    #Adding referee on a new line
    if list_referees_wb.cell(row=irow,column=1).value is None:
        ref_id=int(participant['id'])+200000
        list_referees_wb.cell(row=irow,column=col_for_id).value=ref_id
        list_referees_wb.cell(row=irow,column=col_for_name).value=participant['name']
        list_referees_wb.cell(row=irow,column=col_for_country).value=participant['country']
        list_referees_wb.cell(row=irow,column=col_for_categ).value=participant['category']
        list_referees_wb.cell(row=irow,column=col_for_MC).value=""
        list_referees_wb.cell(row=irow,column=col_for_affiliation).value=participant['affiliation']
        list_referees_wb.cell(row=irow,column=col_for_email).value=participant['email']
        wb_obj.save(file)
        participant['ref_id']=ref_id
        jnf.update_reviewer_map(participant['email'],participant['ref_id'])
        return participant
#def add_user_as_referee(user): 


participant=check_and_add_participant_as_new_referee(args.email[0])
print(participant)
jnf.assign_referee(paper_db_id,participant['db_id'])
print("Check that the participant was correctly assigned!!!")
the_deadline=(datetime.datetime.today()+ datetime.timedelta(days=10)).strftime("%d/%m/%Y")
urlform=jnf.assign_referee_in_file(jnf.find_contrib(the_id=paper_id),participant['ref_id'],participant['name'],the_deadline)
jnf.load_referee_files()
ef.send_email(referee_id=participant['ref_id'],paper_id=paper_id,url=urlform,msgfile='message_referee_invite.txt',deadline=the_deadline)
