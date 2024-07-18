import jacow_nd_func as jnf
import email_func as ef 
import openpyxl
import sys
#import time
#import joblib

sys.exit("Do not run this script unless you want to!")

file = "papers_accepted_for_ToC_final.xlsx"
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
papers_wb = wb_obj.active
irow=1
while (papers_wb.cell(row=irow,column=2).value is not None) and irow<262:
    print(papers_wb.cell(row=irow,column=2).value,papers_wb.cell(row=irow,column=8).value)
    the_paper=jnf.get_paper_info(papers_wb.cell(row=irow,column=2).value,use_cache=True,sleep_before_online=1)
    #print(the_paper)
    print(str(the_paper['revisions'][0]['submitter']['full_name']),str(the_paper['revisions'][0]['submitter']['email']))

    replace_dict={ "name" : str(the_paper['revisions'][0]['submitter']['full_name']) }
    auth_email= str(the_paper['revisions'][0]['submitter']['email'])
    #auth_email= "delerue@lal.in2p3.fr"
    ef.email_file(auth_email,'message_author_volume_published.txt',replace_dict)
    irow=irow+1
        
