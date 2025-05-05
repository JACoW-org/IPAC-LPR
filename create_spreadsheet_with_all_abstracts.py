#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on May 5th 2025

Create a spreadsheet with all abstracts

@author: delerue
"""

import params
import openpyxl

#import sys
#import json
#import requests,json
#from datetime import datetime
#import matplotlib.pyplot as plt
#import numpy as np

#sys.path.append('../jacow-ipac-lpr/')
import users_functions as uf
#import jacow_nd_func as jnf
#import papers_functions as pf
import indico_functions as indf

headers = {'Authorization': f'Bearer {params.api_token}'}

file = "all_abstracts.xlsx"


data = indf.indico_get(f'{params.event_url}/manage/abstracts/abstracts.json',datatype='json',headers=headers)
#print(data.text)
#print(data.json().keys())

'''
if 1==0:
    print(len(data['abstracts']))
    iabstract=53
    print(data['abstracts'][iabstract]['id'])
    print(data['abstracts'][iabstract]['friendly_id'])
    #print(data['abstracts'][iabstract]['submitter'])
    print(data['abstracts'][iabstract].keys())
    print(data['abstracts'][iabstract]['reviews'])
    print(len(data['abstracts'][iabstract]['reviews']))
'''

#print(error_votes_spm)
wb_obj = openpyxl.load_workbook(file) 
summary_sheet=wb_obj["All abstracts"]


fields=[ 'id', 'friendly_id'  , 'MC', 'title' , 'speaker' , 'speaker_gender', 'speaker_seniority' , 'speaker_country' , 'speaker_region' , 'speaker_affiliation' , 'state' ] 



irow=1
for ifield in range(len(fields)):
    summary_sheet.cell(row=irow,column=ifield+1).value=fields[ifield]

for idx in range(len(data['abstracts'])):
    irow=irow+1
    abstract=data['abstracts'][idx]
    abstract['url']=f"{params.event_url}abstracts/{abstract['id']}"
    abstract['speaker']=""
    for person in abstract['persons']:
        if person['is_speaker']:
            #print(person.keys())
            #print(abstract.keys())
            #print(abstract['custom_fields'])
            #print(abstract['custom_fields'][3])
            #print(abstract['custom_fields'][4])
            #print(abstract['custom_fields'][3]['value'])
            #print(person['affiliation'])
            #print(person['affiliation_link'])
            abstract['speaker']=person["first_name"]+" "+person["last_name"]
            abstract['speaker_affiliation']=person['affiliation']            
            abstract['speaker_title']=person['title']
            abstract['MC']=abstract['reviewed_for_tracks'][0]['code']
            for custf in abstract['custom_fields']:
                if custf['id']==165:
                    #print(custf['name'],custf['value'])
                    abstract['speaker_gender']=custf['value']
                if custf['id']==166:
                    #print(custf['name'],custf['value'])
                    abstract['speaker_seniority']=custf['value']
            if person['affiliation_link'] is not None:
                abstract['speaker_country']=person['affiliation_link']['country_name']
                abstract['speaker_region']=uf.get_user_region(person['affiliation_link']['country_code'])
            else:
                abstract['speaker_country']=""
                abstract['speaker_region']=""
    for ifield in range(len(fields)):
        no_gap=False
        no_gap=True
        if no_gap:
            summary_sheet.cell(row=irow,column=ifield+1).value=abstract[fields[ifield]]
        else:
            summary_sheet.cell(row=int(abstract['friendly_id']+1),column=ifield+1).value=abstract[fields[ifield]]

wb_obj.save(file)
