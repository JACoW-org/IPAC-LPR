import jacow_nd_func as jnf
import email_func as ef 
import openpyxl
import time
import joblib
import datetime
import glob,sys


isubmitted=0
isent=0
ireceived=0
papers_accepted= "papers_accepted.xlsx"
n_days_before_resend=500
file = papers_accepted
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
accepted_papers = wb_obj.active

all_validated=glob.glob("author_declaration_forms/validated/*_*")

irow=1
if not "Contrib ID" in accepted_papers.cell(row=irow,column=jnf.PAPER_COL).value:
    print("Error checking ", file)
    exit()
irow=2
while (accepted_papers.cell(row=irow,column=1).value is not None):
    #print('irow', irow)
    contrib_id=accepted_papers.cell(row=irow,column=1).value
    db_id=accepted_papers.cell(row=irow,column=2).value
    title=accepted_papers.cell(row=irow,column=3).value
    submitter_email=accepted_papers.cell(row=irow,column=5).value
    resub_OK=accepted_papers.cell(row=irow,column=9).value
    sub_IoP=accepted_papers.cell(row=irow,column=13).value

    if not(resub_OK=="OK"):
        print("Resub not OK for paper",contrib_id)
        print("resub ", resub_OK)
        print("db_id ", db_id)
        the_paper=jnf.get_paper_info(db_id,sleep_before_online=2)
        if the_paper is None:
            print("Unable to find paper ", db_id)
            exit()
        print(the_paper['state']['name'])
        if the_paper['state']['name'] == 'accepted':
            accepted_papers.cell(row=irow,column=9).value="OK"
            accepted_papers.cell(row=irow,column=8).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb_obj.save(file)
        else:
            print(sub_IoP)
            if not "Withdrawn" in sub_IoP:
                print("Paper state requires investigation")
                exit()
            
        
    if sub_IoP is not None and "Withdrawn" in sub_IoP:
        print("Not submitted to IoP",contrib_id)
        #exit()
    else:
        isubmitted=isubmitted+1
        if accepted_papers.cell(row=irow,column=14).value is not None and (len(str(accepted_papers.cell(row=irow,column=14).value))>3):
            #print("Request already recorded: ",accepted_papers.cell(row=irow,column=14).value)
            #print("checking if reply received ",contrib_id)
            list=glob.glob("author_declaration_forms/validated/"+str(contrib_id)+"_*")
            if (len(list)>0):
                #print(list)
                if accepted_papers.cell(row=irow,column=15).value is None or (len(str(accepted_papers.cell(row=irow,column=15).value))<3):
                    substitution_dict={}
                    substitution_dict['paper_id']=str(contrib_id)
        
                    ef.email_file(submitter_email,'message_author_paper_accepted_thanks_author_declaration_form.txt',replace_dict=substitution_dict)
                    accepted_papers.cell(row=irow,column=15).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    wb_obj.save(file)
                    isent=isent+1
                #else:
                    #print("Reply already received")
                print("contrib_id",contrib_id," file:",list,"  data: ",str(accepted_papers.cell(row=irow,column=15).value),len(str(accepted_papers.cell(row=irow,column=15).value)))
                ireceived=ireceived+1
                index_file=all_validated.index(list[0])
                #print("index: ", index_file)
                #print(len(all_validated))
                all_validated.pop(index_file)
                #print(all_validated)
                #print(len(all_validated))
            else:
                #reply not yet received
                date_request=accepted_papers.cell(row=irow,column=14).value 
                if (accepted_papers.cell(row=irow,column=16).value is not None) and len(str(accepted_papers.cell(row=irow,column=16).value))>3:
                    date_request=accepted_papers.cell(row=irow,column=16).value
                print("Request sent ",date_request, " for paper ", title , " by ", submitter_email)
                date_status=datetime.datetime.now()-datetime.datetime.strptime(str(date_request),"%Y-%m-%d %H:%M:%S")
                print("days ago: ", date_status)
                if date_status.days>=n_days_before_resend:
                    print("send reminder")                    
                    the_paper=jnf.get_paper_info(db_id,sleep_before_online=2)
                    substitution_dict={}
                    substitution_dict['title']=the_paper['contribution']['title']
                    substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
                    substitution_dict['url_paper']="https://indico.jacow.org/event/41/papers/"+str(the_paper['contribution']['id'])+"/"
        
                    ef.email_file(submitter_email,'message_author_paper_accepted_ask_author_declaration_form_reminder.txt',replace_dict=substitution_dict)
                    accepted_papers.cell(row=irow,column=16).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    wb_obj.save(file)
                    isent=isent+1
        else:
            print("Author email", submitter_email)
            the_paper=jnf.get_paper_info(db_id,sleep_before_online=2)
                  
            substitution_dict={}
            substitution_dict['title']=the_paper['contribution']['title']
            substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
            substitution_dict['url_paper']="https://indico.jacow.org/event/41/papers/"+str(the_paper['contribution']['id'])+"/"
        
            ef.email_file(submitter_email,'message_author_paper_accepted_ask_author_declaration_form.txt',replace_dict=substitution_dict)
            accepted_papers.cell(row=irow,column=14).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb_obj.save(file)
            isent=isent+1
    if isent>10:
        print(isent, "emails sent, exiting")
        sys.exit()

    irow=irow+1

print('isent',isent)
print('isubmitted',isubmitted)    
print('ireceived',ireceived)
    
print('all_validated',all_validated)
print(len(all_validated))
