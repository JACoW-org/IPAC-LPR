#Reads additional referees file
#Nicolas Delerue 3/2023

import jacow_nd_func as jnf
import openpyxl

file_all = "all_referees.xlsx"
wb_obj_all = openpyxl.load_workbook(file_all) 
# Read the active sheet:
list_all_referees_wb = wb_obj_all.active

jnf.load_referee_files()

file = "registrations_20230323.xlsx"
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
list_registrants_wb = wb_obj.active
    
irow=1
while (list_all_referees_wb.cell(row=irow,column=1).value is not None):
    categ=list_all_referees_wb.cell(row=irow,column=jnf.col_for_categ)
    if categ.value is None:
        print("None on row",irow)
        email=list_all_referees_wb.cell(row=irow,column=jnf.col_for_email).value
        print('email', email)
        jrow=1
        while (list_registrants_wb.cell(row=jrow,column=3).value is not None) and not email in list_registrants_wb.cell(row=jrow,column=3).value:
            jrow = jrow +1
        print(list_registrants_wb.cell(row=jrow,column=3).value)
        print(list_registrants_wb.cell(row=jrow,column=6).value)
        if list_registrants_wb.cell(row=jrow,column=6).value is not None:
            list_all_referees_wb.cell(row=irow,column=jnf.col_for_categ).value=list_registrants_wb.cell(row=jrow,column=6).value
        else:
            list_all_referees_wb.cell(row=irow,column=jnf.col_for_categ).value=""
    irow=irow+1
wb_obj_all.save(file_all)
