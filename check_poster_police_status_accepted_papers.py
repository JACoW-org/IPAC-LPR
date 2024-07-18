import jacow_nd_func as jnf
import email_func as ef 
import openpyxl
import time
import joblib
import datetime

import requests

#joblib.load(all_entries,"poster_police_results.jlib")
all_entries=joblib.load("poster_police_results.jlib")

papers_accepted= "papers_accepted.xlsx"
file = papers_accepted
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
accepted_papers = wb_obj.active

irow=1
if not "Contrib ID" in accepted_papers.cell(row=irow,column=jnf.PAPER_COL).value:
    print("Error checking ", file)
    exit()


pdf_fail=0    
poster_police_fail=0    
author_fail=0
n_notif=0
max_notif=50
irow=2
while (accepted_papers.cell(row=irow,column=1).value is not None):
    contrib_id=accepted_papers.cell(row=irow,column=1).value
    db_id=accepted_papers.cell(row=irow,column=2).value
    title=accepted_papers.cell(row=irow,column=3).value
    submitter_email=accepted_papers.cell(row=irow,column=5).value
    
    #find paper with given contrib_id
    the_entry=None
    for entry in all_entries:
        if int(entry['contrib_id'])==int(contrib_id):
            the_entry=entry
            
    if the_entry is not None:
            print('contrib_id',contrib_id,the_entry['contrib_id'])
            print('db_id',db_id,the_entry['db_id'])
            print('title',title,the_entry['title'])
            print('status',the_entry['status'])
            print('pdf_status',the_entry['pdf_status'])
            print('police_status',the_entry['police_status'])
            print('author_status',the_entry['author_status'])
            print('author_present_status',the_entry['author_present_status'])
            if not the_entry['status'] == "Accepted":
                print("Not yet accepted")
                #exit()
                pdf_fail=pdf_fail+1
            elif not the_entry['pdf_status'] == "OK":
                print("XXX PDF status")
                time.sleep(1)
                pdf_fail=pdf_fail+1
            elif len(the_entry['police_status'])>0 and not the_entry['police_status']=="OK":
                print('XXX => police_status',the_entry['police_status'])
                time.sleep(1)
                poster_police_fail=poster_police_fail+1
                #exit()
                '''    
                elif (not the_entry['author_status'] == "OK") or (not the_entry['author_present_status']  == "OK"):
                print("XXX authors status")
                time.sleep(1)
                author_fail=author_fail+1
                '''                
                #else:
                #Now we accept all papers
            if 1==1:
                contrib=jnf.find_contrib(the_entry['contrib_id'])
                print(contrib['track'])
                the_paper=jnf.get_paper_info(db_id,use_cache=False,sleep_before_online=0.1)
                print('state', the_paper['state']['name'])
                if the_paper['state']['name']=="Accepted" or the_paper['state']['name']=="accepted": 
                    print("Paper accepted...")
                    substitution_dict={}
                    substitution_dict['title']=title
                    substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
                    substitution_dict['url_paper']="https://indico.jacow.org/event/41/papers/"+str(the_paper['contribution']['id'])+"/"
                    print("Re opening submission for ", substitution_dict['url_paper'])
                    ef.email_file(submitter_email,"message_author_paper_accepted_resubmit_IoP_format.txt",replace_dict=substitution_dict)
                    comment="Hello,\nWe have now received the submission information for publication in IoP proceedings. We will use this indico server to collect the papers. Please upload your paper here as a PDF file with the format required by the IoP as explained in the guidelines posted at https://publishingsupport.iopscience.iop.org/author-guidelines-for-conference-proceedings/ .\nThank you in advance,\nNicolas Delerue"
                    jnf.reopen_paper(db_id)
                    jnf.judge_paper(db_id, "To be corrected" , comment)
                    accepted_papers.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    wb_obj.save(file)
                    n_notif=n_notif+1
                    if n_notif>max_notif:
                        print("Maximum number of notifications reached!")
                        exit()
                elif the_paper['state']['name']=="submitted":
                    print("*** Paper resubmitted "+ str(the_paper['contribution']['friendly_id']) +" ***")
                    print("https://indico.jacow.org/event/41/papers/"+str(db_id)+"/")
                elif the_paper['state']['name']=="to_be_corrected":
                    print("Skipping to_be_corrected state")
                else:
                    print("Other state", the_paper['state']['name'])
                    exit()
    else:
        print("Accepted paper ", entry['contrib_id']," not found")
        exit()
    irow=irow+1
    

print('pdf_fail',pdf_fail)
print('poster_police_fail',poster_police_fail)
print('author_fail',author_fail)

    
