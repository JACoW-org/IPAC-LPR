#!/usr/local/bin/python3.9

# Find info about a paper
# Nicolas Delerue, 3/2023

import jacow_nd_func as jnf
import argparse
import email_func as ef 
import openpyxl
import datetime

def double_printf(reviews,field):
    print(field,eval('reviews[0]'+field),'\t',eval('reviews[1]'+field))
def double_print(reviews,field):
    print(eval('reviews[0]'+field),'\t',eval('reviews[1]'+field))
    
parser = argparse.ArgumentParser()
parser.add_argument("--id", nargs=1, help="The ID of the paper to be checked.")
parser.add_argument("--db_id", nargs=1, help="The db_id of the paper to be checked.")
parser.add_argument("--mail-spb", nargs=1, help="Prepares the email for the SPB, parameter: decision suggested.")
parser.add_argument("--revision", nargs=1, help="Specifies the revision number.")
parser.add_argument("--reject-comment", action="store_true", help="Rejects the last comment that was submitted.")
parser.add_argument("--withdraw", action="store_true", help="Paper withdrawn")
parser.add_argument("--email", action="store_true", help="Find email address of primary author.")
parser.add_argument("--emails", action="store_true", help="Find email addresses of all authors.")
parser.add_argument("--record-review", action="store_true", help="Records the review in the review file.")
parser.add_argument("--notify-authors", action="store_true", help="Notifies the authors of the SPB decision.")
parser.add_argument("--add-review", nargs=1 , help="Bypass on second round to get review that was left as comment")
parser.add_argument("--record-resubmit",action="store_true", help="Record the resubmission date for papers where this was not done automatically")
parser.add_argument("--iop-accept",action="store_true", help="Record the acceptance in IoP format")
parser.add_argument("--iop-reject",action="store_true", help="Informs the author that his paper is not compliant with IoP format.")
parser.parse_args()
args = parser.parse_args()

#print('args',args)
#exit()

authors_email=None
authors_emails=[]
    
if args.db_id is not None:
    print("Looking for info about paper db_id="+args.db_id[0])
    contrib=jnf.find_contrib(the_db_id=args.db_id[0])
    contrib_id=contrib['id']
    paper_db_id=contrib['db_id']

    
if args.id is not None:
    print("Looking for info about paper id="+args.id[0])
    contrib=jnf.find_contrib(the_id=args.id[0])
    #print('contrib',contrib)
    contrib_id=contrib['id']
    paper_db_id=contrib['db_id']

refs=jnf.check_referees_for_paper(contrib_id,show_row=True)
print('Referees',refs)

if args.withdraw:
    print(refs[1])
    for the_ref in refs[1]:
        print('Referee is ',the_ref[0])
        replace_dict={ "name" : the_ref[1] , 'paper_db_id' :  str(paper_db_id) , 'title': contrib['title'] }
        ref_data=jnf.referee_data_file_get_line(the_ref[0])
        recipients=ref_data['email']
        ef.email_file(recipients,"message_referee_paper_withdrawn.txt",replace_dict)
        ref_db_id=jnf.get_referee_db_id_from_id(the_ref[0])
        jnf.unassign_referee(contrib['db_id'],ref_db_id)
        jnf.update_referee_in_file(contrib,the_ref[0],"withdrawn")
    exit()
    
print('paper_db_id',paper_db_id)
print("url: https://indico.jacow.org/event/41/papers/"+str(paper_db_id)+"/")
paper=jnf.get_paper_info(paper_db_id,use_cache=False)
if paper is None:
    print("Paper seems not to exist!!! get_paper_info returned None...")
    exit()
#else:
    #print("Paper ", paper_db_id,"is not None")
#print(paper)
#print(" ")

#if args.mail_spb:
#find MC coordinators
MCfile=open("MC_coordinators_emails.txt","r")
lines=MCfile.readlines()
MCfile.close()
mc_coordinators=[]
for line in lines:
    if line[0:3] == contrib['track'][0:3]:
        mc_coordinators=line.split(";")[1:3]
if len(mc_coordinators)<2:
    print("MC Coordinators not found")
    print('Contrib ', contrib['track'])
    exit()
#print('mc_coordinators',mc_coordinators)
review_msg="Review complete for paper "+str(contrib_id)+"\n"
msg_to_authors="Review complete for paper "+str(contrib_id)+" - "+str(contrib['title'])+"\n"
emails="Nicolas Delerue <nicolas.delerue@ijclab.in2p3.fr>,Peter McIntosh - STFC UKRI <peter.mcintosh@stfc.ac.uk>,Frank Zimmermann <Frank.Zimmermann@cern.ch>,"+mc_coordinators[0].split(",")[1]+","+mc_coordinators[1].split(",")[1]+"\n"
print('Title',contrib['title'])
print('MC',contrib['track'])
print("Code: ",paper['contribution']['code'])
print("url: https://indico.jacow.org/event/41/papers/"+str(paper['contribution']['id'])+"/")
print('Submitter: ',paper['last_revision']['submitter']['full_name'],' ',paper['last_revision']['submitter']['email'])

geodict=jnf.get_countries_from_contrib(contrib)
print('Continent ', geodict['continent'])
print('Countries ',geodict['countries'])

if len(paper['last_revision']['submitter']['email'])>5:
    authors_emails.append(paper['last_revision']['submitter']['email'])
    if authors_email is None:
        authors_email=paper['last_revision']['submitter']['email']
for type in [ 'speakers' , 'primaryauthors' , 'coauthors']:
    for speak in contrib[type]:
        print(type,speak['fullName'],speak['affiliation'])
        if args.email or args.emails:
            user=jnf.search_user(last_name=speak['last_name'],first_name=speak['first_name'])
            if user is not None:
                print(user)
                if user['email'] is not None and len(user['email'])>5:
                    authors_emails.append(user['email'])
                    if authors_email is None:
                        authors_email=user['email']
                        print("Author's email: ", authors_email)
print(" ")

[n_declined,list_declined]=jnf.check_paper_for_declined(contrib_id)
print('Declined by ',list_declined)


#print("Paper", paper)
#print("Paper", paper.keys())
print("Paper state", paper['state']['name'])
print("Revisions", len(paper['revisions']))
revision_number=paper['last_revision']['number']

if args.revision is not None:
    revision_number=int(args.revision[0])
    print("Using revision ",revision_number,' instead of ',paper['last_revision']['number'])
print("Timeline: ",len(paper['revisions'][revision_number-1]['timeline']))
comments_nd=[]
the_reviews=[]
for timel in paper['revisions'][revision_number-1]['timeline']:
    if 'user' in timel.keys():
        email=timel['user']['email']
        print('type: ',timel['timeline_item_type'],' by ', email)
    else:
        print('type: ',timel['timeline_item_type'])
        email=""
    if timel['timeline_item_type']=="comment":
        if len(timel['text'])<100:
            print(timel['text'])
        else:
            print(timel['text'][0:100]+"...")
            
        if "delerue" in email:
            comments_nd.append(timel['text'])
    elif timel['timeline_item_type']=="review":
        the_reviews.append(timel)
        #the_ref=jnf.get_referee_by_email(email)
        print(timel['proposed_action']['name'])

        if args.record_review:
            the_ref=jnf.find_referee_by_email(timel['user']['email'])
            if len(the_ref)==1:
                the_line=jnf.referee_assignation_file_get_line(contrib_id,the_ref[0])
                if not (the_line['status']=='Review received'):
                    print("Review was not recorded for paper ", contrib_id,"  ref ",the_ref[0]) 
                    #print(the_line)
                    jnf.notification_review_received(the_ref[0],contrib_id,timel['proposed_action']['name'],timel['created_dt'],paper['contribution']['title'])

        
review_msg=review_msg+"Dear SPB members,\nThe paper below has been reviewed by "+str(len(the_reviews))+" referees.\n\n"
msg_to_authors=msg_to_authors+"Dear author,\n\nThank you for submitting the paper below to the IPAC'23 Light Peer Review.\n"
review_msg=review_msg+"Title: "+str(contrib['title'])+"\n"
msg_to_authors=msg_to_authors+"Title: "+str(contrib['title'])+"\n"
review_msg=review_msg+"Main Classification: "+str(contrib['track'])+"\n"
msg_to_authors=msg_to_authors+"Main Classification: "+str(contrib['track'])+"\n"
review_msg=review_msg+"Paper code: "+str(paper['contribution']['code'])+"\n"
msg_to_authors=msg_to_authors+"Paper code: "+str(paper['contribution']['code'])+"\n"
review_msg=review_msg+"Paper url: https://indico.jacow.org/event/41/papers/"+str(paper['contribution']['id'])+"/"+"\n"
msg_to_authors=msg_to_authors+"Paper url: https://indico.jacow.org/event/41/papers/"+str(paper['contribution']['id'])+"/"+"\n"
review_msg=review_msg+'Submitter: '+str(paper['last_revision']['submitter']['full_name'])+' '+str(paper['last_revision']['submitter']['email'])+"\n\n"

if args.reject_comment:
    comment_id=-1
    if not paper['revisions'][revision_number-1]['timeline'][comment_id]['timeline_item_type'] == 'comment' :
        print('Last item in the timeline is not a comment')
        exit()
    else:
        email=paper['revisions'][revision_number-1]['timeline'][comment_id]['user']['email']
        print('Item ',comment_id,'in the timeline is a comment by ',email)
        the_ref=jnf.get_referee_by_email(email)
        print(the_ref)
        ef.send_email(the_ref[5],contrib_id,msgfile='message_thank_you_comment.txt',url=None,deadline=None)

if jnf.check_paper_in_double_reviews(contrib_id):
    double_review=jnf.get_paper_in_double_reviews(contrib_id)
    print("Paper is already in the double review file.")
    print("double_review['decision_1st_round']",double_review['date_notification'],double_review['decision_1st_round'])
    print("double_review['suggestion_spb_2nd_round']",double_review['date_suggestion_spb_2nd_round'], double_review['suggestion_spb_2nd_round'])
    
    if double_review['suggestion_spb_2nd_round'] is not None and len(double_review['suggestion_spb_2nd_round'])>5:
        print("A 2nd round decision has been suggested for this paper: ",double_review['suggestion_spb_2nd_round'])        
    elif double_review['decision_1st_round'] is not None and len(double_review['decision_1st_round'])>4:
        
        print("Decision on the 1st round was ",double_review['decision_1st_round'])
    else:
        print("Author not yet notified suggestion is ",double_review['suggestion'])
    if double_review['revision'] is None:
        print('Please check paper revision in file')
        exit()
    else:
        #print("double_review['revision']",double_review['revision'])
        #print("paper['last_revision']['number']",paper['last_revision']['number'])
        if int(double_review['revision'])==int(paper['revisions'][revision_number-1]['number']):
            print("Current revision is the 1st round revision.")
        else:
            print("New revision submitted since the end of the 1st round.")
            if args.record_resubmit:
                if int(double_review['revision'])<int(paper['last_revision']['number']):
                    print("Recording resubmission")                    
                    new_data={}
                    new_data['date_resubmitted']=paper['last_revision']['submitted_dt']
                    jnf.write_paper_in_double_reviews(paper['contribution']['friendly_id'],new_data)
                    exit()
            if args.mail_spb is None:
                if double_review['decision_1st_round']=="Minor revisions (no second round)":

                    print("Review reports on first round:")
                    print(jnf.paper_review_report_txt(paper,for_spb=True,revision_number=double_review['revision']))
                    print("(Review reports on first round)")
                elif double_review['decision_1st_round']=="To be corrected":
                    print("Review reports on this round:")
                    #print(args.add_review)
                    print(jnf.paper_review_report_txt(paper,for_spb=True,revision_number=paper['revisions'][revision_number-1]['number'],add_review=args.add_review,require_two_reviews=False))
                    print("(Review reports on this round)")
                    
                else:
                    print("Paper resubmitted after decision was ", double_review['decision_1st_round'])
                    #exit()
                print("double_review['decision_1st_round']",double_review['decision_1st_round'])
                if double_review['suggestion_spb_2nd_round'] is not None and len(double_review['suggestion_spb_2nd_round'])>5:
                    print("SPB notified on ", double_review['date_suggestion_spb_2nd_round'],double_review['suggestion_spb_2nd_round'])
                    if double_review['date_authors_notified_2nd_round'] is not None and (len(double_review['date_authors_notified_2nd_round'])>5):
                        print("Authors notified on ", double_review['date_authors_notified_2nd_round'])
            else:
                print("Mail spb 2nd round")
                #spb notification after 2nd round
                if not args.mail_spb[0] in ["accept", "reject" ]:
                    print("Argument to --mail-spb ", args.mail_spb[0] , " not understood (2nd round)")
                    exit()
                elif double_review['suggestion_spb_2nd_round'] is not None and len(double_review['suggestion_spb_2nd_round'])>5:
                    print("SPB already notified on second round on ",double_review['date_suggestion_spb_2nd_round'],double_review['suggestion_spb_2nd_round'])
                    print("Paper state", paper['state']['name'])
                    if double_review['date_authors_notified_2nd_round'] is not None and (len(double_review['date_authors_notified_2nd_round'])>5):
                        print("Authors notified on ", double_review['date_authors_notified_2nd_round'])
                    exit()
                else:
                    new_data={}
                    new_data['suggestion_spb_2nd_round']=args.mail_spb[0]
                    new_data['date_suggestion_spb_2nd_round']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    substitution_dict={}
                    substitution_dict['title']=contrib['title']
                    substitution_dict['paper_id']=paper['contribution']['friendly_id']
                    substitution_dict['url_papier']="https://indico.jacow.org/event/41/papers/"+str(paper['contribution']['id'])+"/"
                    
                    substitution_dict['submitter']=str(paper['revisions'][revision_number-1]['submitter']['full_name'])+' '+str(paper['revisions'][revision_number-1]['submitter']['email'])
                    substitution_dict['proposed_action']=args.mail_spb[0]
                    
                    if double_review['decision_1st_round']=="Minor revisions (no second round)":
                        substitution_dict['reviewers_comments']=jnf.paper_review_report_txt(paper,for_spb=True,revision_number=double_review['revision'])
                        jnf.assign_mc_judges(paper['contribution']['id'],mc_coordinators)
                        print("Mails to spb desactivated")
                        #ef.email_file(emails,'message_spb_paper_resubmitted_after_minor_corrections.txt',replace_dict=substitution_dict)
                        jnf.write_paper_in_double_reviews(paper['contribution']['friendly_id'],new_data)
                    elif double_review['decision_1st_round']=="To be corrected":

                        substitution_dict['reviewers_comments']=jnf.paper_review_report_txt(paper,for_spb=True,revision_number=paper['revisions'][revision_number-1]['number'],add_review=args.add_review)
                        jnf.assign_mc_judges(paper['contribution']['id'],mc_coordinators)
                        print("Mails to spb desactivated")
                        #ef.email_file(emails,'message_spb_paper_resubmitted_after_corrections.txt',replace_dict=substitution_dict)
                        jnf.write_paper_in_double_reviews(paper['contribution']['friendly_id'],new_data)
                    else:
                        print("Not ready for this case of decision")
                        exit()

                
if len(the_reviews)>=2 or args.notify_authors:
    prab_fwd=False
    #double_printf(the_reviews,"['created_dt']")
    #print(len(the_reviews), " reviews have been received for this paper")
    if len(the_reviews)>0:
        review_msg=review_msg+"#####\nReviewers' reports:\n"
        msg_to_authors=msg_to_authors+"\nPlease find below the reviewers' reports.\n"

    if len(the_reviews)>=2:
        review_msg=review_msg+"\n"+str(len(the_reviews))+" reviews have been received for this paper.\n"
        for irat in range(len(the_reviews[0]['ratings'])):
            #print(the_reviews[0]['ratings'][irat]['question']['title'])
            review_msg=review_msg+"\n"+the_reviews[0]['ratings'][irat]['question']['title']+"\n"
            if "PR-AB" in the_reviews[0]['ratings'][irat]['question']['title']:
                prab_fwd=True
                for irev in range(0,len(the_reviews)):
                    if not the_reviews[irev]['ratings'][irat]['value']:
                        prab_fwd=False
            if not "PR-AB" in the_reviews[0]['ratings'][irat]['question']['title'] and not "editors"  in the_reviews[0]['ratings'][irat]['question']['title'] and not "originality"  in the_reviews[0]['ratings'][irat]['question']['title'] and not "can be accepted"  in the_reviews[0]['ratings'][irat]['question']['title'] and not "accepted after corrections?"  in the_reviews[0]['ratings'][irat]['question']['title']:
                msg_to_authors=msg_to_authors+"\n"+the_reviews[0]['ratings'][irat]['question']['title']+"\n"            
                for irev in range(0,len(the_reviews)):
                    msg_to_authors=msg_to_authors+"=> Reviewer "+str(irev+1)+": "+str(the_reviews[irev]['ratings'][irat]['value'])+"\n"
                msg_to_authors=msg_to_authors+"\n"
            if not "Comment" in the_reviews[0]['ratings'][irat]['question']['title']:
                for irev in range(0,len(the_reviews)):
                    #print(the_reviews[irev]['ratings'][irat]['value'])
                    #double_print(the_reviews,"['ratings']["+str(irat)+"]['value']")
                    review_msg=review_msg+str(the_reviews[irev]['ratings'][irat]['value'])+"     "
                review_msg=review_msg+"\n"
            else:
                for irev in range(0,len(the_reviews)):
                    #print("Ref "+str(irev+1)+":",the_reviews[irev]['ratings'][irat]['value'])
                    review_msg=review_msg+"=> Reviewer "+str(irev+1)+": "+str(the_reviews[irev]['ratings'][irat]['value'])+"\n"
    for irev in range(0,len(the_reviews)):
        #print("Ref "+str(irev+1)+": ",the_reviews[irev]['comment'])
        if len(the_reviews[irev]['comment'])>2:
            msg_to_authors=msg_to_authors+"Referee "+str(irev+1)+": "+str(the_reviews[irev]['comment'])+"\n"
    msg_to_authors=msg_to_authors+"\n"
    #print(" ")
    #print("Propositions: ")
    if len(the_reviews)>0:
        review_msg=review_msg+"####\nReviewers' proposed action:\n"
    for irev in range(0,len(the_reviews)):
        review_msg=review_msg+"Reviewer "+str(irev+1)+":  "+the_reviews[irev]['proposed_action']['name']+"\n"
        #print(the_reviews[irev]['proposed_action']['name'])
    if prab_fwd:
        print("*** Forward to PR-AB ***")
    #print(review_msg)
    #jnf.check_double_reviews(contrib_id)
    #print(jnf.check_paper_in_double_reviews(contrib_id))
    if jnf.check_paper_in_double_reviews(contrib_id):
        double_review_data=jnf.get_paper_in_double_reviews(contrib_id)
        print("double_review_data['date_notification']",double_review_data['date_notification'])
        print("double_review_data['date_resubmitted']",double_review_data['date_resubmitted'])
        print("double_review_data['suggestion_spb_2nd_round']",double_review_data['suggestion_spb_2nd_round'])
        print("double_review_data['date_suggestion_spb_2nd_round']",double_review_data['date_suggestion_spb_2nd_round'])
        print("double_review_data['date_authors_notified_2nd_round']",double_review_data['date_authors_notified_2nd_round'])
        if args.notify_authors and (
            double_review_data['date_notification'] is None or len(double_review_data['date_notification'])<5
            or (
                double_review_data['date_resubmitted'] is not None and len(double_review_data['date_resubmitted'])>5
                and (double_review_data['date_authors_notified_2nd_round'] is None or len(double_review_data['date_authors_notified_2nd_round'])<5)
                )
            ):
            is_second_round=double_review_data['date_resubmitted'] is not None and len(double_review_data['date_resubmitted'])>5
            if is_second_round:
                if not paper['state']['name'] == "submitted":
                    print("Paper state", paper['state']['name'])
                    print("Authors can only be notified when the paper state is submitted")
                    exit()
            print("Notify_authors")
            msg_to_authors=msg_to_authors+"Based on the reviewers report the Scientific publication board took the following decision:\n"
            paper_data={}
            paper_comment=''
            if not is_second_round and (double_review['suggestion'] == "To be corrected" or double_review['suggestion'] == "Minor revisions (no second round)"):
                #msg_to_authors=msg_to_authors+"The paper is to be corrected and resubmitted after the reviewers' comments are taken into account.\nPlease correct the paper and resubmit it within 5 days of receiving this message.\n"
                msg_to_authors=msg_to_authors+"The paper is to be corrected and resubmitted after the reviewers' comments are taken into account.\nPlease correct the paper and resubmit it before Thursday 11th May noon.\n"
                msg_to_authors=msg_to_authors+"\nThank you in advance,\n\nNicolas Delerue\non behalf of the IPAC'23 LPR SPB"
                paper_data['decision_1st_round']=double_review['suggestion']
                paper_comment='Paper to be corrected.\n'+msg_to_authors
            elif (double_review['suggestion'] == "accept") or (is_second_round and double_review['suggestion_spb_2nd_round'] == "accept"):
                msg_to_authors=msg_to_authors+"Congratulations, the paper is accepted. No more actions are required from you at this stage. You may be contacted by the Jacow editors regarding the Jacow version of your paper and we will contact you after IPAC'23 to ask you to submit your paper to the IoP Proceedings website.\n"
                msg_to_authors=msg_to_authors+"\nBest wishes,\n\nNicolas Delerue\non behalf of the IPAC'23 LPR SPB"
                if is_second_round:
                    paper_data['decision_spb_2nd_round']=double_review['suggestion_spb_2nd_round']
                    paper_data['date_authors_notified_2nd_round']=datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                else:
                    paper_data['decision_1st_round']=double_review['suggestion']
                    paper_comment='Paper accepted.'
                jnf.record_accepted_paper(contrib_id)
            elif double_review['suggestion'] == "reject" or (is_second_round and double_review['suggestion_spb_2nd_round'] == "reject"):
                msg_to_authors=msg_to_authors+"Unfortunately, after considering carefully the reviewers reports, we decided that this paper can not be accepted.\n"
                msg_to_authors=msg_to_authors+"\nBest wishes,\n\nNicolas Delerue\non behalf of the IPAC'23 LPR SPB"
                if is_second_round:
                    paper_data['decision_spb_2nd_round']=double_review['suggestion_spb_2nd_round']
                    paper_data['date_authors_notified_2nd_round']=datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                else:
                    paper_data['decision_1st_round']=double_review['suggestion']
                    paper_comment='Paper rejected. See email for details.'
                jnf.record_rejected_paper(contrib_id)
            else:
                print("double_review['suggestion']", double_review['suggestion'])
                print('is_second_round',is_second_round)
                print("Not yet handled")
                exit()
            #print('msg_to_authors')
            #print(msg_to_authors)
            if authors_email is not None:
                paper_data['date_notification']=datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                if double_review['revision'] == "None":
                    paper_data['revision']=revision_number
                #print(paper_data)
                jnf.write_paper_in_double_reviews(contrib_id,paper_data)
                #unassign judges
                the_judges=[]
                #Peter
                the_judges.append(857)
                #jnf.assign_judge(paper_id,jnf.get_referee_db_id_from_id(427))
                #Frank
                the_judges.append(1000)
                #jnf.assign_judge(paper_id,jnf.get_referee_db_id_from_id(845))
                for mc_coord in mc_coordinators:
                    mc_user=jnf.search_user(email=mc_coord.split(",")[1])
                    if mc_user is None:
                        print('unable to find',mc_coord)
                    else:
                        print('mc_user',mc_user['email'],mc_user['id'])
                        #jnf.assign_judge(paper_id,mc_user['id'])
                        if not "delerue" in mc_user['email']: 
                            the_judges.append(mc_user['id'])
                jnf.unassign_judge(paper_db_id,the_judges)

                if is_second_round:
                    jnf.judge_paper(paper_db_id, paper_data['decision_spb_2nd_round'],paper_comment)
                else:
                    jnf.judge_paper(paper_db_id, paper_data['decision_1st_round'],paper_comment)
                print('authors_email',authors_email)
                ef.email_txt(authors_email,msg_to_authors)
                if len(comments_nd)>0:
                    print("*** Comments by ND ***\n",comments_nd)
            else:
                print("No email for this paper")
                exit()        

    elif args.mail_spb is not None:
        #spb notification for paper not yet in double review file            
        if not args.mail_spb[0] in ["accept", "minor", "tbc" , "reject" ]:
            print("Argument to --mail-spb ", args.mail_spb[0] , " not understood")
            exit()
        else:
            if args.mail_spb[0]=="tbc":
                args.mail_spb[0]="To be corrected"
            if args.mail_spb[0]=="minor":
                args.mail_spb[0]="Minor revisions (no second round)"
            review_msg=review_msg+"\n\n\nThe proposed action is: "+args.mail_spb[0]+"\n\nPlease let me know within 24h if you disagree with this proposal.\n\nThank you in advance,\n\nNicolas\n"
    
        print("Recording...")
        jnf.assign_mc_judges(paper_db_id,mc_coordinators)
        file = "papers_double_review.xlsx"
        wb_obj = openpyxl.load_workbook(file) 

        # Read the active sheet:
        double_review_wb = wb_obj.active

        irow=1

        if not "Contrib ID" in double_review_wb.cell(row=irow,column=1).value:
            print("Error checking ", file)
            exit()

        #Check where to enter new data
        irow=1
        while (double_review_wb.cell(row=irow,column=1).value is not None):
            irow=irow+1
            #print("Checking row if is available", irow)
    
        double_review_wb.cell(row=irow,column=1).value=contrib_id
        double_review_wb.cell(row=irow,column=2).value=paper_db_id
        double_review_wb.cell(row=irow,column=3).value=contrib['track']
        double_review_wb.cell(row=irow,column=4).value=the_reviews[0]['proposed_action']['name']
        double_review_wb.cell(row=irow,column=5).value=the_reviews[1]['proposed_action']['name']
        double_review_wb.cell(row=irow,column=6).value=args.mail_spb[0]
        if prab_fwd:
            double_review_wb.cell(row=irow,column=7).value="Yes"
        else:
            double_review_wb.cell(row=irow,column=7).value="No"
            
        double_review_wb.cell(row=irow,column=8).value=datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        double_review_wb.cell(row=irow,column=12).value=revision_number

        #double_review_wb
        wb_obj.save(file)

        ###ef.email_txt("peer-review@ipac23.org",review_msg)                
        ef.email_txt(emails,review_msg)
    else:
        print(review_msg)
        print("Paper is not yet in the double review file.")
    
if args.email:
    print("Author's email address: ", authors_email)
if args.emails:
    print("Authors' emails: ", authors_emails)

if args.iop_accept or args.iop_reject:
    print("IoP accept",contrib_id, paper_db_id)
    print(authors_email)

    substitution_dict={}
    substitution_dict['title']=contrib['title']
    substitution_dict['paper_id']=paper['contribution']['friendly_id']
    substitution_dict['url_paper']="https://indico.jacow.org/event/41/papers/"+str(paper['contribution']['id'])+"/"
                    
    substitution_dict['submitter']=str(paper['revisions'][revision_number-1]['submitter']['full_name'])+' '+str(paper['revisions'][revision_number-1]['submitter']['email'])
    
    file = jnf.papers_accepted
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    accepted_papers = wb_obj.active

    irow=2
    while not int(accepted_papers.cell(row=irow,column=1).value) == int(contrib_id) and accepted_papers.cell(row=irow,column=1).value is not None:
        irow=irow+1

    if accepted_papers.cell(row=irow,column=1).value is not None:
        print('irow',irow)
    else:
        print("Unable to find paper in accepted_papers")
        exit()

    if args.iop_accept:
        paper_comment="Thank you for resubmitting your paper in the IoP format"
        jnf.judge_paper(paper_db_id, "accept" ,paper_comment)                    
        ef.email_file(authors_email,'message_author_paper_accepted_resubmit_IoP_format_Thank_You.txt',replace_dict=substitution_dict)
        accepted_papers.cell(row=irow,column=9).value="OK"
    else:
        paper_comment="The paper is not compliant with the IoP guidelines. Please see https://publishingsupport.iopscience.iop.org/author-guidelines-for-conference-proceedings/ "
        jnf.judge_paper(paper_db_id, "To be corrected" ,paper_comment)                    
        ef.email_file(authors_email,'message_author_paper_accepted_resubmit_IoP_format_check.txt',replace_dict=substitution_dict)
        accepted_papers.cell(row=irow,column=9).value="Rejected"
        
    accepted_papers.cell(row=irow,column=8).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb_obj.save(file)

