#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 28 16:40:34 2025

Checks reviews on abstracts

@author: delerue
"""

import params
import openpyxl

#import sys
#import json
#import requests,json
#from datetime import datetime
#import matplotlib.pyplot as plt
#import numpy as np

#sys.path.append('../jacow-ipac-lpr/')
import users_functions as uf
#import jacow_nd_func as jnf
#import papers_functions as pf
import indico_functions as indf

headers = {'Authorization': f'Bearer {params.api_token}'}

file = "votes_summary.xlsx"


data = indf.indico_get(f'{params.event_url}/manage/abstracts/abstracts.json',datatype='json',headers=headers)
#print(data.text)
#print(data.json().keys())

'''
if 1==0:
    print(len(data['abstracts']))
    iabstract=53
    print(data['abstracts'][iabstract]['id'])
    print(data['abstracts'][iabstract]['friendly_id'])
    #print(data['abstracts'][iabstract]['submitter'])
    print(data['abstracts'][iabstract].keys())
    print(data['abstracts'][iabstract]['reviews'])
    print(len(data['abstracts'][iabstract]['reviews']))
'''

all_votes_spm={}
all_votes_abstract={}
error_votes_spm={}


print("Checking votes")
for idx in range(len(data['abstracts'])):
    abstract=data['abstracts'][idx]
    if (len(abstract['reviews'])>0):
        for review in abstract['reviews']:
            #print("review ","id",abstract['id'],abstract['friendly_id'],abstract['state'],f"{params.event_url}abstracts/{abstract['id']}/")
            #theroles=uf.has_role_codes(userid=review['user']['id'])
            #print("    user:",review['user']['id'],review['user']['full_name'])
            #print("    ratings:",review['ratings'])
            if review['ratings'][0]['value'] is None and review['ratings'][1]['value'] is None:
                pass
            elif review['ratings'][0]['value'] is False and review['ratings'][1]['value'] is False:
                pass
            elif review['ratings'][0]['value'] is True and review['ratings'][1]['value'] is True:
                print("Warning: Double True vote!",abstract['id'],review['user']['full_name'])
                exit()
            elif abstract['state'] in [ 'duplicate' ,'withdrawn' , 'rejected' ]:
                print("Vote on rejected abstract! ",abstract['id'],"by",review['user']['full_name'])
                if not review['user']['id'] in error_votes_spm.keys():
                    error_votes_spm[review['user']['id']]={}
                    error_votes_spm[review['user']['id']]['full_name']=review['user']['full_name']
                    error_votes_spm[review['user']['id']]['votes']=[]
                error_votes_spm[review['user']['id']]['votes'].append({"id": abstract['id'], "state": abstract['state'], "title": abstract['title']}  )
            elif review['user']['id']==4587:
                #ignoring the votes by Sabrina
                pass
            else:
                if review['user']['full_name']=="Nicolas Delerue" and 0==1:
                    print(review['user']['full_name'])
                    print(abstract['id'])
                    print(abstract['reviewed_for_tracks'][0]['title'][0:3])
                    print(abstract['submitted_for_tracks'][0]['title'][0:3])
                    #print(abstract.keys())
                    #exit()
                if not review['user']['id'] in all_votes_spm.keys():
                    all_votes_spm[review['user']['id']]={}
                    all_votes_spm[review['user']['id']]['full_name']=review['user']['full_name']
                    if "cerf" in  review['user']['full_name'] and 0==1:
                        print(review['user']['full_name'])
                        print(review['user']['id'])
                        exit()
                    for imc in range(1,9):
                        all_votes_spm[review['user']['id']]["MC"+str(imc)]={}
                        all_votes_spm[review['user']['id']]["MC"+str(imc)]['1st']=[]
                        all_votes_spm[review['user']['id']]["MC"+str(imc)]['2nd']=[]                            
                        
                if not abstract['id'] in all_votes_abstract.keys():
                    all_votes_abstract[abstract['id']]={}
                    all_votes_abstract[abstract['id']]['MC']=abstract['reviewed_for_tracks'][0]['title'][0:3]
                    all_votes_abstract[abstract['id']]['1st']=[]
                    all_votes_abstract[abstract['id']]['2nd']=[]
                    all_votes_abstract[abstract['id']]['idx']=idx
                if review['ratings'][0]['value'] is True:
                    all_votes_spm[review['user']['id']][abstract['reviewed_for_tracks'][0]['title'][0:3]]['1st'].append(abstract['id'])                    
                    all_votes_abstract[abstract['id']]['1st'].append(review['user']['id'])
                elif review['ratings'][1]['value'] is True:
                    all_votes_spm[review['user']['id']][abstract['reviewed_for_tracks'][0]['title'][0:3]]['2nd'].append(abstract['id'])                    
                    all_votes_abstract[abstract['id']]['2nd'].append(review['user']['id'])


#print(error_votes_spm)
wb_obj = openpyxl.load_workbook(file) 

full_names=sorted([ [all_votes_spm[thekey]['full_name'] , thekey ] for thekey in all_votes_spm.keys() ])
full_names_list=sorted([ all_votes_spm[thekey]['full_name'] for thekey in all_votes_spm.keys() ])
               
#print(full_names)
#Summarize and write in an excel sheet
summary_sheet=wb_obj["Votes summary"]
iname=0
istartrow=1
for fname in full_names:
    for irow in range(0,6):
        summary_sheet.cell(row=iname*4+istartrow+irow,column=1).value=""
    summary_sheet.cell(row=iname*4+istartrow,column=1).value=fname[0]
    print(fname[0])
    print("     1st choices: ",end='')
    summary_sheet.cell(row=iname*4+1+istartrow,column=2).value="1st choices"
    for imc in range(1,9):
        summary_sheet.cell(row=iname*4+istartrow,column=2+imc).value="MC"+str(imc)
        print(f"MC{imc}: ",len(all_votes_spm[fname[1]]["MC"+str(imc)]['1st']),end=" ")
        summary_sheet.cell(row=iname*4+1+istartrow,column=2+imc).value=len(all_votes_spm[fname[1]]["MC"+str(imc)]['1st'])
    print("\n     2nd choices: ",end='')
    summary_sheet.cell(row=iname*4+2+istartrow,column=2).value="2nd choices"
    for imc in range(1,9):
        print(f"MC{imc}: ",len(all_votes_spm[fname[1]]["MC"+str(imc)]['2nd']),end=" ")
        summary_sheet.cell(row=iname*4+2+istartrow,column=2+imc).value=len(all_votes_spm[fname[1]]["MC"+str(imc)]['2nd'])
    print("")
    iname=iname+1
    

#print(full_names_list)
SPC_members=uf.get_SPC_members(params.event_id)
print(SPC_members)
istartrow=iname*4+istartrow+4
iname=1
for member in SPC_members:
    if not member in full_names_list:
        print(member,"has not yet voted")
        summary_sheet.cell(row=istartrow+iname,column=1).value=str(member)+": No vote yet"
        iname=iname+1
    
for imc in range(1,9):
    #if imc>1:
    #    mc_sheet=wb_obj.create_sheet(title="MC"+str(imc))
    mc_sheet=wb_obj["MC"+str(imc)]
    print(f"MC{imc}")
    cols=[ 'id' , 'friendly_id', 'title' , 'speaker', 'affiliation' , 'url' ]
    
    irow=1
    icol=1
    mc_sheet.cell(row=irow,column=icol).value="Score"
    icol=icol+1
    mc_sheet.cell(row=irow,column=icol).value="1st choice"
    icol=icol+1
    mc_sheet.cell(row=irow,column=icol).value="2nd choice"
    icol=icol+1
    for col in cols:
        mc_sheet.cell(row=irow,column=icol).value=col
        icol=icol+1
    
    ranked=sorted([ [ 2*len(all_votes_abstract[theid]['1st'])+len(all_votes_abstract[theid]['2nd']),  theid] for theid in all_votes_abstract.keys() if all_votes_abstract[theid]['MC'] == "MC"+str(imc) ], reverse=True)
    irow=2
    for therank in ranked:
        abstract=data['abstracts'][all_votes_abstract[therank[1]]['idx']]
        abstract['url']=f"{params.event_url}abstracts/{abstract['id']}"
        abstract['speaker']=""
        for person in abstract['persons']:
            if person['is_speaker']:
                abstract['speaker']=abstract['speaker']+person["first_name"]+" "+person["last_name"]
                abstract['affiliation']=person['affiliation']
        #abstract['MC']=abstract['reviewed_for_tracks'][0]['code']
        icol=1
        mc_sheet.cell(row=irow,column=icol).value=therank[0]
        icol=icol+1
        mc_sheet.cell(row=irow,column=icol).value=len(all_votes_abstract[therank[1]]['1st'])
        icol=icol+1
        mc_sheet.cell(row=irow,column=icol).value=len(all_votes_abstract[therank[1]]['2nd'])
        icol=icol+1
        for col in cols:
            mc_sheet.cell(row=irow,column=icol).value=abstract[col]
            icol=icol+1
        irow=irow+1
        
wb_obj.save(file)
