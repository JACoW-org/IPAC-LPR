#!/usr/local/bin/python3

# Extracts the evaluation report for all accepted papers from indico
# Nicolas Delerue, 9/2023

import jacow_nd_func as jnf
import email_func as ef
import openpyxl
#import time
#import joblib
import datetime
import sys

isubmitted=0
papers_accepted= "papers_accepted.xlsx"
file = papers_accepted
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
accepted_papers = wb_obj.active

irow=1
if not "Contrib ID" in accepted_papers.cell(row=irow,column=jnf.PAPER_COL).value:
    print("Error checking ", file)
    sys.exit()
irow=2
while accepted_papers.cell(row=irow,column=1).value is not None:
    contrib_id=accepted_papers.cell(row=irow,column=1).value
    db_id=accepted_papers.cell(row=irow,column=2).value
    title=accepted_papers.cell(row=irow,column=3).value
    submitter_email=accepted_papers.cell(row=irow,column=5).value
    resub_OK=accepted_papers.cell(row=irow,column=9).value
    sub_IoP=accepted_papers.cell(row=irow,column=13).value

    if not(resub_OK=="OK"):
        print("Resub not OK for paper",contrib_id)
        print("resub ", resub_OK)
        print("db_id ", db_id)
        the_paper=jnf.get_paper_info(db_id,sleep_before_online=2)
        if the_paper is None:
            print("Unable to find paper ", db_id)
            sys.exit()
        print(the_paper['state']['name'])
        if the_paper['state']['name'] == 'accepted':
            accepted_papers.cell(row=irow,column=9).value="OK"
            accepted_papers.cell(row=irow,column=8).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb_obj.save(file)
        else:
            print(sub_IoP)
            if not "Withdrawn" in sub_IoP:
                print("Paper state requires investigation")
                sys.exit()
            
        
    if sub_IoP is not None and "Withdrawn" in sub_IoP:
        print("Not submitted to IoP",contrib_id)
        #exit()
    else:
        isubmitted=isubmitted+1
        the_paper=jnf.get_paper_info(db_id,sleep_before_online=2)

        #report
        print(" ======================================= ")
        print("ID: ", the_paper['contribution']['id'])
        print("Title: ", the_paper['contribution']['title'])
        print("Submitter:  ", the_paper['revisions'][0]['submitter']['full_name'])
        print("affiliation: ", the_paper['revisions'][0]['submitter']['affiliation'])
        print("email address: ", the_paper['revisions'][0]['submitter']['email'])

        for irevision in range(0,len(the_paper['revisions'])):
            print("Revision ",irevision+1)
            the_rev=the_paper['revisions'][irevision]
            print("Date submitted:", the_rev['submitted_dt'])
            #print(len( the_rev['timeline']))
            for itimeline in range(len( the_rev['timeline'])):
                spacer="    "
                the_item=the_rev['timeline'][itimeline]
                #print(the_item['timeline_item_type'])
                if the_item['timeline_item_type']=="review":
                    print(spacer,"Review:")
                    spacer=spacer+"   "
                    print(spacer,"Created: ", the_item['created_dt'])
                    print(spacer,"Reviewer: ", the_item['user']['full_name'],"  ",the_item['user']['affiliation'],"  ",the_item['user']['affiliation_meta']['country_name'])
                    spacer=spacer+"   "
                    for irat in range(len(the_item['ratings'])):
                        the_rat=the_item['ratings'][irat]
                        if not "PRAB" in the_rat['question']['title']:
                            print(spacer, the_rat['question']['title'])
                            print(spacer, the_rat['value'])
                if the_item['timeline_item_type']=="judgment":
                    #print(the_rev.keys())
                    #print(the_rev['judge']['full_name'])
                    print(spacer,"Judgement")
                    spacer=spacer+"   "
                    if "IoP proceedings" in the_rev['judgment_comment']:
                        print(spacer,"Accepted - file to be resubmitted in the correct format")
                    elif "IoP guidelines" in the_rev['judgment_comment']:
                        print(spacer,"File format to be corrected")                        
                    else:
                        print(spacer,"Decision ", the_rev['state'])
                        print(spacer,"Decision comment", the_rev['judgment_comment'])
                    
        print("   ")
        #sys.exit()
        '''        
        print(the_paper)
        print(the_paper.keys())
        print(len(the_paper['revisions']))
        print(the_paper['revisions'][0].keys())
        print(the_paper['revisions'][0]['timeline'])
        print(len(the_paper['revisions'][0]['timeline']))
        print(the_paper['revisions'][0]['timeline'][0])
        print(the_paper['revisions'][0]['timeline'][0]['user'])
        print(the_paper['revisions'][0]['timeline'][0]['timeline_item_type'])
        print(the_paper['revisions'][0]['timeline'][0].keys())
        sys.exit()
        '''
        
    irow=irow+1
    
print('isubmitted',isubmitted)    

    
