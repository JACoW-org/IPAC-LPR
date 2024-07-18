import jacow_nd_func as jnf
import email_func as ef 
import openpyxl
import time
import joblib
import datetime
import os

import pdfquery
import lxml.etree as etree
import requests

papers_accepted= "papers_accepted.xlsx"
file = papers_accepted
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
accepted_papers = wb_obj.active

irow=1
if not "Contrib ID" in accepted_papers.cell(row=irow,column=jnf.PAPER_COL).value:
    print("Error checking ", file)
    exit()

n_accepted=0
n_tbc=0
list_tbc = []
n_submitted=0

n_notif=0
max_notif=10
age_resend_notif=3
irow=2
while (accepted_papers.cell(row=irow,column=1).value is not None):
    print('irow',irow)
    contrib_id=accepted_papers.cell(row=irow,column=1).value
    db_id=accepted_papers.cell(row=irow,column=2).value
    title=accepted_papers.cell(row=irow,column=3).value
    submitter_email=accepted_papers.cell(row=irow,column=5).value
    
            
    print('contrib_id',contrib_id)
    print('db_id',db_id)
    print('title',title)
    contrib=jnf.find_contrib(contrib_id)
    print(contrib['track'])
    #add track to accepted papers file
    if 1==1:
        accepted_papers.cell(row=irow,column=10).value=contrib['track'][0:3]
        accepted_papers.cell(row=irow,column=11).value=contrib['track'][4:7]
        wb_obj.save(file)
        
    the_paper=jnf.get_paper_info(db_id,use_cache=False,sleep_before_online=2)
    print('state', the_paper['state']['name'])
    substitution_dict={}
    substitution_dict['title']=title
    substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
    substitution_dict['url_paper']="https://indico.jacow.org/event/41/papers/"+str(the_paper['contribution']['id'])+"/"
    if the_paper['state']['name']=="Accepted" or the_paper['state']['name']=="accepted": 
        print("Paper accepted -- nothing to be done")
        n_accepted=n_accepted+1
    elif the_paper['state']['name']=="submitted":
        n_submitted=n_submitted+1
        print("*** Paper resubmitted "+ str(the_paper['contribution']['friendly_id']) +" ***")
        print("https://indico.jacow.org/event/41/papers/"+str(db_id)+"/")
        #print('Files submitted',len(the_paper['last_revision']['files']))
        if len(the_paper['last_revision']['files'])>1:
            print("Too many files submitted")
            print("Re opening submission for ", substitution_dict['url_paper'])
            ef.email_file(submitter_email,"message_author_paper_accepted_resubmit_IoP_format_too_many_files.txt",replace_dict=substitution_dict)
            comment="Hello,\nPlease upload only the PDF file for the IoP proccedings.\nThank you in advance,\nNicolas Delerue"
            jnf.reopen_paper(db_id)
            jnf.judge_paper(db_id, "To be corrected" , comment)
            #wb_obj.save(file)
            n_notif=n_notif+1
            if n_notif>max_notif:
                print("Maximum number of notifications reached!")
                exit()
        elif the_paper['last_revision']['spotlight_file'] is None:
            print("Spotlight file is None")
            print(the_paper['last_revision']['files'][0])
            print(the_paper['last_revision']['files'][0]['filename'])
            print(the_paper['last_revision']['files'][0]['filename'][-4:].lower())
            if not the_paper['last_revision']['files'][0]['filename'][-4:].lower()==".pdf":
                print("Not PDF")
            exit()
        else:
            if len(the_paper['last_revision']['timeline'])>0:
                #print(the_paper['last_revision']['timeline'])
                print('Timeline entry:')
                print(the_paper['last_revision']['timeline'][-1]['timeline_item_type'])
                print(the_paper['last_revision']['timeline'][-1]['text'])
                print(the_paper['last_revision']['timeline'][-1]['user']['full_name'])
                if not the_paper['last_revision']['timeline'][-1]['user']['full_name'] == "Nicolas Delerue":
                    #exit()
                    time.sleep(2)
            print('spotlight file',the_paper['last_revision']['spotlight_file'])
            print('content_type',the_paper['last_revision']['spotlight_file']['content_type'])
            print('filename',the_paper['last_revision']['spotlight_file']['filename'])
            print('download_url',the_paper['last_revision']['spotlight_file']['download_url'])
            url= "https://indico.jacow.org/"+ the_paper['last_revision']['spotlight_file']['download_url']
            if 1==1:
                pdfname='/Users/delerue/Downloads/ipac23_papers_accepted/ipac23_paper_'+contrib['track'][0:7].replace(".","_")+'_id_'+str(contrib_id)+'_db_'+str(db_id)+'.pdf'
                if os.path.exists(pdfname):
                    print("File already exists")
                    print('open ',pdfname)
                else:
                    print("File does not exists")
                    file_request = requests.get(url, allow_redirects=True)
                    if not file_request.status_code == 200:
                        print("Error downloading file")
                        exit()
                    else:
                        open(pdfname, 'wb').write(file_request.content)
                    if not os.path.exists(pdfname):
                        print("File is missing")
                    else:
                        print('Downloaded under: open ',pdfname)
                        pdf_file = pdfquery.PDFQuery(pdfname)
                        try:
                            pdf_file.load()
                            print(pdf_file.tree.getroot().attrib['Producer'])
                            print(pdf_file.tree.getroot().attrib.keys())
                            if 'Title' in pdf_file.tree.getroot().attrib.keys():
                                if "Journal of Physics: Conference series" in pdf_file.tree.getroot().attrib['Title']:
                                    print("Title OK")
                                else:
                                    print("??? Title: ", pdf_file.tree.getroot().attrib['Title'])
                            else:
                                print("No Title")
                                #check some keywords that can be the sign of an error
                                for keyword in [ 'bema' , '??' ,  'Bema' , 'xx' , 'XX' ]:
                                    val=pdf_file.pq('LTTextLineHorizontal:contains("'+keyword+'")')
                                    if len(val.text())>0:
                                        print(len(val.text()),val.text(),val)
                                        print("keyword ",keyword,"found... To be checked. Contrib id:",contrib_id)
                                        if int(contrib_id) not in [ 1499 , 1115 , 1020 , 1464 , 1971 , 2035 ]:
                                            exit()
                        except:
                            print("Loading pdf failed")
                            if int(contrib_id) not in [ 1478 , 798 , 2677 ]:
                                exit()
                        time.sleep(8)
                if os.path.exists(pdfname):
                    print('To accept: ./find_paper.py --iop-accept --id',the_paper['contribution']['friendly_id'])
                    #exit()
    elif the_paper['state']['name']=="to_be_corrected":
        n_tbc=n_tbc+1
        list_tbc.append(the_paper['contribution']['friendly_id'])
        print("Paper still in to_be_corrected state")
        print('Date tbc',the_paper['last_revision']['judgment_dt'])
        age_status=datetime.datetime.now()-datetime.datetime.strptime(str(the_paper['last_revision']['judgment_dt'][0:10]),"%Y-%m-%d")
        print('age_status',age_status)
        if age_status.days>age_resend_notif:
            if len(the_paper['last_revision']['timeline'])>0:
                if the_paper['last_revision']['timeline'][-1]['timeline_item_type'] == 'judgment':
                    pass
                elif the_paper['last_revision']['timeline'][-1]['timeline_item_type'] == 'comment':
                    age_status=datetime.datetime.now()-datetime.datetime.strptime(str(the_paper['last_revision']['timeline'][-1]['created_dt'][0:10]),"%Y-%m-%d")
                    print('age_status',age_status)
                else:                    
                    print('Timeline entry:')
                    print(the_paper['last_revision']['timeline'][-1])
                    exit()

            if age_status.days>age_resend_notif:
                substitution_dict={}
                substitution_dict['title']=the_paper['contribution']['title']
                substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
                substitution_dict['url_paper']="https://indico.jacow.org/event/41/papers/"+str(the_paper['contribution']['id'])+"/"
                print("Late resubmission for paper ", the_paper['contribution']['friendly_id'])
                ef.email_file(submitter_email,"message_author_paper_accepted_resubmit_IoP_format_reminder.txt",replace_dict=substitution_dict)
                comment="This is a reminder that you must re-upload your paper in the format requested by IoP. See guidelines at https://publishingsupport.iopscience.iop.org/author-guidelines-for-conference-proceedings/\nPlease upload only the PDF file."
                jnf.comment_paper(db_id, comment)
                #exit()

        #exit()
    else:
        print("Other state", the_paper['state']['name'])
        exit()
    irow=irow+1
    
print('n_accepted',n_accepted)
print('n_tbc',n_tbc)
print('list_tbc',list_tbc)
print('n_submitted',n_submitted)
print('done')
    
