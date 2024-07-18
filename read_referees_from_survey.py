#Reads additional referees file
#Nicolas Delerue 3/2023

import jacow_nd_func as jnf
import openpyxl

file_all = "all_referees.xlsx"
wb_obj_all = openpyxl.load_workbook(file_all) 
# Read the active sheet:
list_all_referees_wb = wb_obj_all.active

jnf.load_referee_files()

irow=1
while (list_all_referees_wb.cell(row=irow,column=1).value is not None):
    irow=irow+1
print('row',irow)

file = "survey_volunteering_LPR_20230321.xlsx"
wb_obj = openpyxl.load_workbook(file) 
# Read the active sheet:
list_referees_wb = wb_obj.active

jrow=1
if not "Submitter Email" in list_referees_wb.cell(row=jrow,column=2).value:
    print("Error checking ", file)
    exit()
    
#Looks for the data line by line
jrow=2
while (list_referees_wb.cell(row=jrow,column=2).value is not None):
    print(list_referees_wb.cell(row=jrow,column=2).value)

    email=list_referees_wb.cell(row=jrow,column=2).value
    ref_info=jnf.get_referee_info_from_email(email)
    #print(ref_info)
    id=100000+int(ref_info['id'])
    name=ref_info['full_name']
    date=list_referees_wb.cell(row=jrow,column=3).value
    MC=list_referees_wb.cell(row=jrow,column=9).value
    affiliation=ref_info['affiliation']
    categ=""
    country=jnf.get_country_from_affiliation(affiliation)

    print(email,date,MC,country,id,name)
    jrow=jrow+1

    list_all_referees_wb.cell(row=irow,column=jnf.col_for_id).value=id
    list_all_referees_wb.cell(row=irow,column=jnf.col_for_name).value=name
    list_all_referees_wb.cell(row=irow,column=jnf.col_for_country).value=country
    list_all_referees_wb.cell(row=irow,column=jnf.col_for_MC).value=MC
    list_all_referees_wb.cell(row=irow,column=jnf.col_for_categ).value=categ
    list_all_referees_wb.cell(row=irow,column=jnf.col_for_email).value=email
    list_all_referees_wb.cell(row=irow,column=jnf.col_for_affiliation).value=affiliation 
    irow=irow+1
    
wb_obj_all.save(file_all)

