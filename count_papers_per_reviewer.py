import jacow_nd_func as jnf
import email_func as ef 
import argparse
import openpyxl
import datetime
import urllib.parse
import requests
import time

if 1==1:
    jnf.load_referee_files()
    
    file = jnf.papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    all_referees_with_papers=[]
    
    irow=2
    while not referees_assignation_wb.cell(row=irow,column=1).value is None:
        ref_id=referees_assignation_wb.cell(row=irow,column=3).value
        ref_name=referees_assignation_wb.cell(row=irow,column=4).value
        print('ref_id', ref_id,ref_name)
        ref_data=jnf.referee_data_file_get_line(ref_id)
        data={}
        print(ref_data)
        if ref_data is None:
            print('no ref data')
            exit()
        print('status',referees_assignation_wb.cell(row=irow,column=6).value)
        if referees_assignation_wb.cell(row=irow,column=6).value == 'Email request sent':
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
            print('sent')
        elif referees_assignation_wb.cell(row=irow,column=6).value == 'Accepted':
            print('papers_accepted')
            if not 'papers_accepted' in ref_data.keys() or ref_data['papers_accepted'] is None:
                data['papers_accepted']=1
            else:
                data['papers_accepted']=ref_data['papers_accepted']+1
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        elif referees_assignation_wb.cell(row=irow,column=6).value == 'Declined':
            print('declined')
            if not 'papers_declined' in ref_data.keys()  or ref_data['papers_declined'] is None:
                data['papers_declined']=1
            else:
                data['papers_declined']=ref_data['papers_declined']+1
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        elif referees_assignation_wb.cell(row=irow,column=6).value == 'Withdrawn':
            print('Withdrawn')
            if not 'papers_withdrawn' in ref_data.keys()  or ref_data['papers_withdrawn'] is None:
                data['papers_withdrawn']=1
            else:
                data['papers_withdrawn']=ref_data['papers_withdrawn']+1
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        elif referees_assignation_wb.cell(row=irow,column=6).value == 'Review received' :
            print('review')
            all_referees_with_papers.append(ref_id)
            if not 'reviewed' in ref_data.keys() or ref_data['papers_reviewed'] is None:
                data['papers_reviewed']=1
            else:
                data['papers_reviewed']=ref_data['papers_reviewed']+1
            if not 'papers_accepted' in ref_data.keys() or ref_data['papers_accepted'] is None:
                data['papers_accepted']=1
            else:
                data['papers_accepted']=ref_data['papers_accepted']+1
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        elif referees_assignation_wb.cell(row=irow,column=6).value == '2nd Reminder sent':
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        elif referees_assignation_wb.cell(row=irow,column=6).value == 'Reminder sent':
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        elif referees_assignation_wb.cell(row=irow,column=6).value == 'Deadline reminder sent':
            if not 'papers_accepted' in ref_data.keys() or ref_data['papers_accepted'] is None:
                data['papers_accepted']=1
            else:
                data['papers_accepted']=ref_data['papers_accepted']+1
            if not 'papers_assigned' in ref_data.keys() or ref_data['papers_assigned'] is None:
                data['papers_assigned']=1
            else:
                data['papers_assigned']=ref_data['papers_assigned']+1
        else:
            print('status not understood')
            exit()
        
        #jnf.referee_data_file_write_line(ref_id,data)
        irow=irow+1
    print('done')
    print('all_referees_with_papers',len(all_referees_with_papers))
    all_referees_with_papers_listed=list(set(all_referees_with_papers))
    print('all_referees_with_papers_listed',len(all_referees_with_papers_listed),all_referees_with_papers_listed)
    emails=[]
    emails_str=""
    max_papers=0
    email_max=''
    for the_ref in all_referees_with_papers_listed:
        email=jnf.get_referee_by_id(the_ref)[1]
        n_papers=jnf.check_papers_for_referee(the_ref)['n_papers']
        if n_papers>max_papers:
            max_papers=n_papers
            email_max=email
        #print(jnf.get_referee_by_id(the_ref))
        #print(jnf.check_papers_for_referee(the_ref))
        print(email,n_papers)
        emails.append(email)
        emails_str=emails_str+","+email
    print(emails)
    print(emails_str)
    print('Max papers= ',max_papers,email_max)
exit()
