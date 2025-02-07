#Set of functions to sort jacow data for LPR
#Nicolas Delerue 3/2023

#export PYTHONPATH=.

import openpyxl
import numpy as np
import requests
import json
import os,time,sys
import random
import joblib
import email_func as ef 
import datetime
import urllib.parse

#Settings
event_id = 41

if not ('api_token' in locals() or 'api_token' in globals()):
    fname="../api_token.txt"
    try:
        fpwd=open(fname)
        api_token=fpwd.readlines()[0]
        fpwd.close()
    except:
        print("Unable to read ",fname)
        print("You need to create a file called api_token containing your api_token in the directory above the one where the LPR scripts are located")
        sys.exit(1)

event_url="https://indico.jacow.org/event/"+str(event_id)+"/"

#print("jacow_nd_func imported")
Requests_timeout=20

col_for_id=-1
col_for_name=-1
col_for_country=-1
col_for_MC=-1
col_for_categ=-1
col_for_email=-1
col_for_affiliation=-1

max_referees=5000
max_per_referee_papers=3

list_referees_wb = None

allMCs=[]
referees_for_all_MCs=[]
referees_for_all_MCs_by_continent=[]

n_referees=0
n_student_referees=0

data_json_contribs=None

papers_referee_assignation = "papers_referee_assignation.xlsx"
allrefdata_fname='all_referees_by_sub_MC.data'

n_days_to_resubmit_after_1st_round=7
days_to_notify_authors_after_1st_round=1

all_actions_todo=[]


def printv(txt):
    print(txt," : ",eval(txt))

#### Referees
def load_referee_files():
    global max_referees
    global list_referees_wb

    global col_for_id
    global col_for_name
    global col_for_country
    global col_for_MC
    global col_for_categ
    global col_for_email
    global col_for_affiliation
    
    #file = "referees-at-registrations.xlsx"
    file = "all_referees.xlsx"
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    list_referees_wb = wb_obj.active
    #print(list_referees_wb['A1'].value)
    #printv('col_for_id')
    for thecol in list_referees_wb.iter_cols(min_col=1, max_col=20, max_row=1):
        for thecell in thecol:
            if thecell.value is not None:
                if "ID" in thecell.value:
                    col_for_id=thecell.col_idx
                if thecell.value=="Name": 
                    col_for_name=thecell.col_idx
                if "Country" in thecell.value:
                    col_for_country=thecell.col_idx
                if "Category" in thecell.value:
                    col_for_categ=thecell.col_idx
                if "classification" in thecell.value:
                    col_for_MC=thecell.col_idx
                if "Email" in thecell.value:
                    col_for_email=thecell.col_idx
                if "Affiliation" in thecell.value:
                    col_for_affiliation=thecell.col_idx
                
    #printv('col_for_id')
    #printv('col_for_name')
    #printv('col_for_country')
    #printv('col_for_MC')
    
def referee_assignation_file_get_line(paper_id,ref_id,start_row=2):
    #find the matching line
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()
    
    irow=start_row
    while not referees_assignation_wb.cell(row=irow,column=1).value is None and not ((int(referees_assignation_wb.cell(row=irow,column=1).value)==int(paper_id)) and (int(referees_assignation_wb.cell(row=irow,column=3).value)==int(ref_id))):
        irow=irow+1
        #print("Checking row if is available", irow)
    if referees_assignation_wb.cell(row=irow,column=1).value is None:
        return None
    elif ((int(referees_assignation_wb.cell(row=irow,column=1).value)==int(paper_id)) and (int(referees_assignation_wb.cell(row=irow,column=3).value)==int(ref_id))):
        reply_data={}
        reply_data['row']=irow
        reply_data['paper_id']=referees_assignation_wb.cell(row=irow,column=1).value
        reply_data['paper_title']=referees_assignation_wb.cell(row=irow,column=2).value
        reply_data['ref_id']=referees_assignation_wb.cell(row=irow,column=3).value
        reply_data['ref_name']=referees_assignation_wb.cell(row=irow,column=4).value
        reply_data['date_initial_message']=referees_assignation_wb.cell(row=irow,column=5).value
        reply_data['status']=referees_assignation_wb.cell(row=irow,column=6).value
        reply_data['last_communication']=referees_assignation_wb.cell(row=irow,column=7).value
        reply_data['deadline']=referees_assignation_wb.cell(row=irow,column=8).value
        return reply_data
    else:
        return None

def get_all_referees_from_file():
    global list_referees_wb 

    if list_referees_wb is None:
        load_referee_files()

    all_refs=[]
    for cell in list_referees_wb["A"][1:]:
        all_refs.append(cell.value)
    #print(all_refs)
    return all_refs
        
def referee_data_file_get_line(ref_id,start_row=2):
    global list_referees_wb 

    if list_referees_wb is None:
        load_referee_files()
        
    irow=1
    if not "ID" in list_referees_wb.cell(row=irow,column=1).value:
        print("Error checking file list_referee")
        exit()
    
    irow=start_row
    while not list_referees_wb.cell(row=irow,column=1).value is None and not (int(list_referees_wb.cell(row=irow,column=1).value)==int(ref_id)):
        irow=irow+1
        #print("Checking row if is available", irow)
    if list_referees_wb.cell(row=irow,column=1).value is None:
        return None
    elif (int(list_referees_wb.cell(row=irow,column=1).value)==int(ref_id)):
        reply_data={}
        reply_data['row']=irow
        reply_data['id']=list_referees_wb.cell(row=irow,column=1).value
        reply_data['name']=list_referees_wb.cell(row=irow,column=2).value
        reply_data['email']=list_referees_wb.cell(row=irow,column=3).value
        reply_data['affiliation']=list_referees_wb.cell(row=irow,column=4).value
        reply_data['country']=list_referees_wb.cell(row=irow,column=5).value
        reply_data['category']=list_referees_wb.cell(row=irow,column=6).value
        reply_data['accepting_papers']=list_referees_wb.cell(row=irow,column=12).value
        reply_data['papers_assigned']=list_referees_wb.cell(row=irow,column=13).value
        reply_data['papers_declined']=list_referees_wb.cell(row=irow,column=14).value
        reply_data['papers_accepted']=list_referees_wb.cell(row=irow,column=15).value
        reply_data['papers_reviewed']=list_referees_wb.cell(row=irow,column=16).value
        if reply_data['papers_assigned'] is None:
            reply_data['papers_assigned']=0
        if reply_data['papers_declined'] is None:
            reply_data['papers_declined']=0
        if reply_data['papers_accepted'] is None:
            reply_data['papers_accepted']=0
        if reply_data['papers_reviewed'] is None:
            reply_data['papers_reviewed']=0
        return reply_data
    else:
        return None

def referee_data_file_write_line(ref_id,data):
    file = "all_referees.xlsx"
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    list_referees_wb = wb_obj.active

    irow=1
    if not "ID" in list_referees_wb.cell(row=irow,column=1).value:
        print("Error checking list_referee")
        exit()
    
    irow=2
    while not list_referees_wb.cell(row=irow,column=1).value is None and not (int(list_referees_wb.cell(row=irow,column=1).value)==int(ref_id)):
        irow=irow+1
        #print("Checking row if is available", irow)
    if list_referees_wb.cell(row=irow,column=1).value is None:
        return None
    elif (int(list_referees_wb.cell(row=irow,column=1).value)==int(ref_id)):
        #print(data.keys())
        if 'accepting_papers' in data.keys():
            list_referees_wb.cell(row=irow,column=12).value=data['accepting_papers']
        if 'papers_assigned' in data.keys():
            list_referees_wb.cell(row=irow,column=13).value=data['papers_assigned']
        if 'papers_declined' in data.keys():
            list_referees_wb.cell(row=irow,column=14).value=data['papers_declined']
        if 'papers_accepted' in data.keys():
            list_referees_wb.cell(row=irow,column=15).value=data['papers_accepted']
        if 'papers_reviewed' in data.keys():
            list_referees_wb.cell(row=irow,column=16).value=data['papers_reviewed']
        wb_obj.save(file)
        print("Referee data file written")
        load_referee_files()
        return True
    else:
        return None


def get_referee_by_email(email):
    return get_referee_by_xx(email,'email')
    
def get_referee_dict_by_email(email):
    return get_referee_dict_by_xx(email,'email')
    
def get_referee_dict_by_id(referee_id):
    return get_referee_dict_by_xx(referee_id,'id')
    
def get_referee_by_id(referee_id):
    return get_referee_by_xx(referee_id,'id')
    
def get_referee_by_xx(val, typecol):
    ref_dict=get_referee_dict_by_xx(val, typecol)
    if ref_dict is not None:
        return [ ref_dict['name'], ref_dict['email'], ref_dict['affiliation'], ref_dict['country'], ref_dict['MC'] , ref_dict['id'] ]
    else:
        print("get_referee returned nothing for ",typecol," ", val)
        return None
    
def get_referee_dict_by_xx(val, typecol):
    global max_referees
    global list_referees_wb

    global col_for_id
    global col_for_name
    global col_for_country
    global col_for_MC
    global col_for_email
    global col_for_categ
    global col_for_affiliation

    ref_dict={}

    if col_for_id == -1:
        load_referee_files()
        
    if typecol=='id':
        val=int(val)
        the_col=col_for_id
    elif typecol=='email':
        the_col=col_for_email
    else:
        print('unknow typecol',typecol)
        the_col=col_for_id
        exit()
    if list_referees_wb is None:
        load_referee_files()

    id_found=False
    for irow in range(2,max_referees):
        thecell_xx=list_referees_wb.cell(row=irow,column=the_col)
        if thecell_xx.value==val:
            #print('found')
            #printv('col_for_name')
            thecell_id=list_referees_wb.cell(row=irow,column=col_for_id)
            thecell_name=list_referees_wb.cell(row=irow,column=col_for_name)
            #print('Referee ', referee_id,' is: ',thecell_name.value)
            vals=[]
            vals.append(list_referees_wb.cell(row=irow,column=col_for_name).value)
            ref_dict['name']=list_referees_wb.cell(row=irow,column=col_for_name).value
            vals.append(list_referees_wb.cell(row=irow,column=col_for_email).value)
            ref_dict['email']=list_referees_wb.cell(row=irow,column=col_for_email).value
            vals.append(list_referees_wb.cell(row=irow,column=col_for_affiliation).value)
            ref_dict['affiliation']=list_referees_wb.cell(row=irow,column=col_for_affiliation).value
            vals.append(list_referees_wb.cell(row=irow,column=col_for_country).value)
            ref_dict['country']=list_referees_wb.cell(row=irow,column=col_for_country).value
            vals.append(list_referees_wb.cell(row=irow,column=col_for_MC).value)
            ref_dict['MC']=list_referees_wb.cell(row=irow,column=col_for_MC).value
            vals.append(list_referees_wb.cell(row=irow,column=col_for_id).value)
            ref_dict['id']=list_referees_wb.cell(row=irow,column=col_for_id).value
            ref_dict['categ']=list_referees_wb.cell(row=irow,column=col_for_categ).value
            return  ref_dict
            id_found=True
    if not id_found:
        print("Referee with id ", val ," was not found")
        return None
        exit()

        
def get_referee_db_id_from_id(referee_id):
    vals=get_referee_by_id(referee_id)
    print('The referee with id ',referee_id,' has email ',vals[1])
    return get_referee_db_id_from_email(vals[1])
    
def get_referee_db_id_from_email(email):
    referee_data=get_referee_info_from_email(email)    
    #print('referee_data',referee_data)
    #print("referee_data['id']",referee_data['id'])
    #print("referee_data['avatar_url']".referee_data['avatar_url'])
    if referee_data['id'] is not None:
        return referee_data['id']
    #elif referee_data['avatar_url'] is not None:
    #    print(referee_data['avatar_url'].split("/")[2])
    #    exit()
    else:
        print(referee_data)
        print("Unable to find referee id for ",email)
        exit()
          
def get_referee_info_from_email(email):
    headers = {'Authorization': f'Bearer {api_token}'}
    data = requests.get(f'https://indico.jacow.org/user/search/?email='+email, headers=headers)
    #print(data)
    #print(data.json()['users'][0])    
    return data.json()['users'][0]

def assign_referee(paper_db_id,referee_db_id):
    print("assigning referee", referee_db_id, " to paper ", paper_db_id)
    headers = {'Authorization': f'Bearer {api_token}'}
    payload = {'contribution_id': [ paper_db_id ], 'user_id': [ referee_db_id ] }
    data = requests.post(f'{event_url}manage/papers/assignment-list/assign/content_reviewer', headers=headers, data=payload)
    print(data)
    if not (data.status_code == 200):
        exit()

def unassign_referee(paper_id,referee_db_id):
    #print("db_id received by unassign_referee", referee_db_id)
    #print("db_id received by unassign_referee", type(referee_db_id))
    referee_db_id=str(referee_db_id)
    headers = {'Authorization': f'Bearer {api_token}'}
    payload = {'contribution_id': [ paper_id ], 'user_id': [ referee_db_id ] }
    data = requests.post(f'{event_url}manage/papers/assignment-list/unassign/content_reviewer', headers=headers, data=payload)
    print(data)
    if not (data.status_code == 200):
        exit()
    #print("Check paper ", paper_id ," for referee unassigned")


def assign_referee_in_file(contrib,ref_id,ref_name,the_deadline):
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    #Check where to enter new data
    irow=1
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        irow=irow+1
        #print("Checking row if is available", irow)
    
    referees_assignation_wb.cell(row=irow,column=1).value=contrib['id']
    referees_assignation_wb.cell(row=irow,column=2).value=contrib['title']
    referees_assignation_wb.cell(row=irow,column=3).value=ref_id
    referees_assignation_wb.cell(row=irow,column=4).value=ref_name
    referees_assignation_wb.cell(row=irow,column=5).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    referees_assignation_wb.cell(row=irow,column=6).value="Email request sent"
    referees_assignation_wb.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    referees_assignation_wb.cell(row=irow,column=8).value=the_deadline

    #referees_assignation_wb
    wb_obj.save(file)
    entry_code=str(((int(ref_id)*10000)+int(contrib['id']))*7)

    urlassign="http://nicolas.delerue.org/ipac23/assign.php?"
    urlassign=urlassign+urllib.parse.urlencode({'key': str(entry_code), 'title': contrib['title'], 'name' : ref_name })
    #print('url',urlassign)

    data = requests.get(urlassign)
    if not data.status_code == 200:
        print("Error submitting the assignememnt: wrong response code")
    else:
        #print(data.text)
        if not data.text[-2:] == "OK":
            print("Error submitting the assignememnt: wrong reply from server")
        
        urlform="http://nicolas.delerue.org/ipac23/referee_acceptance_form.php?id="+entry_code
        return urlform

def update_referee_in_file(contrib,ref_id,update_type):
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    #Check where to enter new data
    irow=2
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None) and not ((int(referees_assignation_wb.cell(row=irow,column=1).value) == int(contrib['id'])) and (int(referees_assignation_wb.cell(row=irow,column=3).value) == int(ref_id))):
        irow=irow+1
        #print("Checking row if is available", irow)

    if ((int(referees_assignation_wb.cell(row=irow,column=1).value) == int(contrib['id'])) and (int(referees_assignation_wb.cell(row=irow,column=3).value) == int(ref_id))):
        #record the update
        if update_type=='withdrawn':
            referees_assignation_wb.cell(row=irow,column=6).value="Withdrawn"
        else:
            print("Update not understood")
        referees_assignation_wb.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    #referees_assignation_wb
    wb_obj.save(file)

def assign_mc_judges(paper_db_id,mc_coordinators):
        the_judges=[]
        #Peter
        the_judges.append(857)
        #jnf.assign_judge(paper_id,jnf.get_referee_db_id_from_id(427))
        #Frank
        the_judges.append(1000)
        #jnf.assign_judge(paper_id,jnf.get_referee_db_id_from_id(845))
        for mc_coord in mc_coordinators:
            mc_user=search_user(email=mc_coord.split(",")[1])
            print('mc_coord',mc_coord,'mc_user',mc_user)
            if mc_user is None:
                print('unable to find',mc_coord)
            else:
                print('mc_user',mc_user['email'],mc_user['id'])
                #jnf.assign_judge(paper_id,mc_user['id'])
                the_judges.append(mc_user['id'])
        assign_judge(paper_db_id,the_judges)

    
def assign_judge(paper_db_id,referee_db_id):
    global Requests_timeout
    if type(referee_db_id) == list:
        #print('is list')
        list_referee_db_id=referee_db_id
    else:
        list_referee_db_id=[referee_db_id]
    print("assigning judge", referee_db_id, " to paper ", paper_db_id)
    headers = {'Authorization': f'Bearer {api_token}'}
    payload = {'contribution_id': [ paper_db_id ], 'user_id': [ referee_db_id ] }
    data = requests.post(f'{event_url}manage/papers/assignment-list/assign/judge', headers=headers, data=payload,timeout=Requests_timeout)
    print(data)
    print(len(data.text))
    if not (data.status_code == 200):
        print("Assigning judge returned status code: ", data.status_code)
        exit()
    if len(data.text)<5000:
        print("Assigning judge returned too short reply ",len(data.text))
        exit()

    
def unassign_judge(paper_id,referee_db_id):
    global Requests_timeout
    if type(referee_db_id) == list:
        #print('is list')
        list_referee_db_id=referee_db_id
    else:
        list_referee_db_id=[referee_db_id]
    print("unassigning judge", referee_db_id, " to paper ", paper_id)
    headers = {'Authorization': f'Bearer {api_token}'}
    payload = {'contribution_id': [ paper_id ], 'user_id': [ referee_db_id ] }
    data = requests.post(f'{event_url}manage/papers/assignment-list/unassign/judge', headers=headers, data=payload,timeout=Requests_timeout)
    print(data)
    print(len(data.text))
    if not (data.status_code == 200):
        print("Unassigning judge returned status code: ", data.status_code)
        exit()
    if len(data.text)<5000:
        print("Unassigning judge returned too short reply ",len(data.text))
        exit()
    

def find_referee_by_name(referee_name):
    global max_referees
    global list_referees_wb

    global col_for_id
    global col_for_name
    
    if col_for_id == -1:
        load_referee_files()
        
    if list_referees_wb is None:
        load_referee_files()
    id_found=False
    retval=[]
    for irow in range(2,max_referees):
        thecell_name=list_referees_wb.cell(row=irow,column=col_for_name)
        if str(referee_name).strip().lower() in str(thecell_name.value).strip().lower():
            #print('found')
            #printv('col_for_name')
            thecell_id=list_referees_wb.cell(row=irow,column=col_for_id)
            print('Referee ', referee_name,' has id: ',thecell_id.value)
            id_found=True
            retval.append(thecell_id.value)
    if not id_found:
        #print("Referee with name ", referee_name ," was not found")
        #retval=None
        pass
    return retval
    
def find_referee_by_email(referee_email):
    global max_referees
    global list_referees_wb

    global col_for_id
    global col_for_email
    
    if col_for_id == -1:
        load_referee_files()
        
    if list_referees_wb is None:
        load_referee_files()
    id_found=False
    retval=[]
    for irow in range(2,max_referees):
        thecell_email=list_referees_wb.cell(row=irow,column=col_for_email)
        if str(referee_email).strip().lower() in str(thecell_email.value).strip().lower():
            #print('found')
            #printv('col_for_name')
            thecell_id=list_referees_wb.cell(row=irow,column=col_for_id)
            #print('Referee ', referee_email,' has id: ',thecell_id.value)
            id_found=True
            retval.append(thecell_id.value)
    if not id_found:
        #print("Referee with name ", referee_name ," was not found")
        #retval=None
        pass
    return retval
    
def referee_action(paper_id,the_ref_id,action):
    contrib=find_contrib(the_id=paper_id)
    #print('contrib', contrib)
    paper_db_id=contrib['db_id']
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()
    
    ref_data=referee_data_file_get_line(the_ref_id)
    data={}
    #Looks for the line where to enter the data
    irow=1
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None) and not ((str(referees_assignation_wb.cell(row=irow,column=1).value)==str(paper_id)) and (str(referees_assignation_wb.cell(row=irow,column=3).value)==str(the_ref_id))):
        irow=irow+1

    if ((str(referees_assignation_wb.cell(row=irow,column=1).value)==str(paper_id)) and (str(referees_assignation_wb.cell(row=irow,column=3).value)==str(the_ref_id))):
        print("Found matching entry on row",irow)

    if action=='accept':
        referees_assignation_wb.cell(row=irow,column=6).value="Accepted"
        ef.send_email(referee_id=the_ref_id,paper_id=paper_id,msgfile='message_thank_you_accept.txt')
        referees_assignation_wb.cell(row=irow,column=9).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data['papers_accepted']=ref_data['papers_accepted']+1
        referee_data_file_write_line(the_ref_id,data)
    elif action=='decline':
        referees_assignation_wb.cell(row=irow,column=6).value="Declined"
        ef.send_email(referee_id=the_ref_id,paper_id=paper_id,msgfile='message_thank_you_decline.txt')
        referees_assignation_wb.cell(row=irow,column=9).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ref_db_id=get_referee_db_id_from_id(the_ref_id)
        #print('db_id',ref_db_id)
        unassign_referee(paper_db_id,ref_db_id)
        #print("Check that unassignement worked")
        data={}
        data['papers_declined']=ref_data['papers_declined']+1
        referee_data_file_write_line(the_ref_id,data)

    elif action=='unavailable' or action=='overdue':
        record_referee_unavailable(the_ref_id)
        referees_assignation_wb.cell(row=irow,column=6).value="Declined"
        if action=='unavailable':
            ef.send_email(referee_id=the_ref_id,paper_id=paper_id,msgfile='message_thank_you_unavailable.txt')
        else:
            print("Recording overdue")
            record_referee_overdue(the_ref_id)            
        referees_assignation_wb.cell(row=irow,column=9).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        unassign_referee(paper_db_id,get_referee_db_id_from_id(the_ref_id))
        data={}
        data['papers_declined']=ref_data['papers_declined']+1
        data['accepting_papers']='unavailable'
        referee_data_file_write_line(the_ref_id,data)
        
    referees_assignation_wb.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    #referees_assignation_wb
    wb_obj.save(file)
    print("Referee assignation wb written")

def load_referees_unavailable():
    unavailable=[]
    fname="referees_unavailable.txt"
    fdata=open(fname,"r")
    for line in fdata.readlines():
        #print(line)
        if (len(line)>2):
            unavailable.append(int(line.split(";")[0].strip()))
    fdata.close()
    #print(unavailable)
    return unavailable
    
def record_referee_unavailable(referee_id):
    ref=get_referee_by_id(int(referee_id))
    fname="referees_unavailable.txt"
    fdata=open(fname,"a")
    fdata.write(str(referee_id)+";"+ref[0]+"\n")
    fdata.close()

def record_referee_overdue(referee_id):
    ref=get_referee_by_id(int(referee_id))
    fname="referees_overdue.txt"
    fdata=open(fname,"a")
    fdata.write(str(referee_id)+";"+ref[0]+"\n")
    fdata.close()

def can_referee_get_additional_paper(ref_id):
    ref_data=referee_data_file_get_line(ref_id)
    unavailable_referees=load_referees_unavailable()
    if( int(ref_id) in unavailable_referees):
        return False
    categ=ref_data['category']
    if categ is not None and ("Student" in categ):
        return False
    retval=check_papers_for_referee(ref_id)
    n_papers=retval['n_papers']
    accepting_extra_papers=ref_data['accepting_papers']
    if((n_papers>=max_per_referee_papers) and not (accepting_extra_papers=='yes')):
        return False
    if ( int(ref_id)>200000) and not (accepting_extra_papers=='yes') :
        return False
    if n_papers>0:
        hist=check_history_for_referee(ref_id)
        for paper_hist in hist:
            #print('hist',paper_hist)
            if paper_hist['status'] in [ 'Email request sent' , "Reminder sent" , "2nd Reminder sent" ]:
                return False
            else:
                if not accepting_extra_papers=='yes':
                    age_status=datetime.datetime.now()-datetime.datetime.strptime(str(paper_hist['date_status']),"%Y-%m-%d %H:%M:%S")
                    #print('age_status',age_status,paper_hist['date_status'])
                    #exit()
                    if age_status.days<2:
                        return False
    return True

def remove_authors_from_list(contrib,the_list):
    the_list_names=[]
    for auth in the_list:
        if type(auth[2]) is str:
            the_list_names.append(auth[2].lower())
        else:
            the_list_names.append(auth[2]['name'].lower())
    for the_type in [ 'speakers' , 'primaryauthors' , 'coauthors' ]: 
        for speak in contrib[the_type]:
            speak_full_name=speak['first_name'].lower()+" "+speak['last_name'].lower()
            #print(speak_full_name)
            if speak_full_name in the_list_names:
                #print("author in referees",speak_full_name,the_list_names)
                idx=the_list_names.index(speak_full_name)
                #print(idx)
                the_list_names=the_list_names[0:idx]+the_list_names[idx+1:]
                the_list=the_list[0:idx]+the_list[idx+1:]
                #print(the_list_names)
                #exit()
    return the_list

def find_reviewer_for_MC(MC=None,sub=None,excepted=None):
    global allrefdata_fname
    [the_sub_MC_list,all_referees_by_sub_MC,all_referees_by_sub_MC_coauthors]=joblib.load(allrefdata_fname)
    unavailable_referees=load_referees_unavailable()
    refs_dict={}            
    can_get=[]
    can_get_names=[]
    can_not_get=[]
    additionals=[]
    if sub is not None and not(len(sub))==3:
        print("Length of subMC should be 3!!!")
        exit()
    for subMC in the_sub_MC_list:
        if (not (subMC == excepted)) and (((MC is not None) and (subMC[0:len(MC)]==MC)) or ((sub is not None) and (subMC[4:7]==sub))):
            idx=the_sub_MC_list.index(subMC)
            print("*** ", subMC, "***")
            print('authors',all_referees_by_sub_MC[idx])
            print('co-authors',all_referees_by_sub_MC_coauthors[idx])
            for auth in list(set(all_referees_by_sub_MC_coauthors[idx]+all_referees_by_sub_MC[idx])):
                if auth in all_referees_by_sub_MC[idx]:
                    authtxt="author"
                else:
                    authtxt="co-author"
                #print('auth', auth,type(auth))
                if type(auth) is int:
                    if not( int(auth) in unavailable_referees):
                        if can_referee_get_additional_paper(int(auth)):
                            the_ref=get_referee_dict_by_id(auth)

                            can_get.append([ authtxt , int(auth), the_ref['name'] ] )
                        else:
                            can_not_get.append([ authtxt , int(auth)] )
                else:
                    participant=find_participant_by_email(auth)
                    #print("Additional ",authtxt,auth,participant)
                    additionals.append( [authtxt,auth,participant])

    refs_dict={}
    refs_dict['can_get']=can_get
    refs_dict['can_not_get']=can_not_get
    refs_dict['additionals']=additionals
    return refs_dict
    
def assign_referees_to_MC():
    global max_referees
    global list_referees_wb

    global col_for_id
    global col_for_name
    global col_for_country
    global col_for_MC

    global allMCs
    global referees_for_all_MCs
    global referees_for_all_MCs_by_continent

    global n_referees
    global n_student_referees
    global max_per_referee_papers

    
    allMCs=[]
    referees_for_all_MCs=[]
    referees_for_all_MCs_by_continent=[]
    n_referees=0

    #Loads data for the "split" MCs (MC6 and MC7)
    fextraname='split_MC_referees.data'
    extra_refs=joblib.load(fextraname)
    
    if list_referees_wb is None:
        load_referee_files()        

    #We read the excel file of the referees line by line    
    for irow in range(2,max_referees):
        thecell_id=list_referees_wb.cell(row=irow,column=col_for_id)

        if thecell_id.value is not None:
            ref_id=str(thecell_id.value)
            if can_referee_get_additional_paper(ref_id):
                n_referees=n_referees+1
                #thecell_name=list_referees_wb.cell(row=irow,column=col_for_name)
                thecell_country=list_referees_wb.cell(row=irow,column=col_for_country)
                the_continent=get_continent(thecell_country.value)
                thecell_MC=list_referees_wb.cell(row=irow,column=col_for_MC)
                #print('value ',thecell_id.value,thecell_name.value,thecell_country.value,thecell_MC.value)
                
                #If it is None, it is all
                if thecell_MC.value is None:
                    thecell_MC.value="MC1;MC2;MC3;MC4;MC5"
                #check if this referee has extra contributions in split MCs
                idx=-1
                if thecell_id.value in [ ref[0] for ref in extra_refs]:
                    idx=([ ref[0] for ref in extra_refs]).index(thecell_id.value)
                    for subMc in extra_refs[idx][1]:
                        thecell_MC.value=thecell_MC.value+";"+str(subMc[0])
                    #print(thecell_MC.value)
                for MC_val in thecell_MC.value.split(";"):
                    theMC=MC_val.split(":")[0]
                    if not theMC in allMCs:
                        allMCs.append(theMC)
                        referees_for_all_MCs.append([])
                        referees_for_all_MCs_by_continent.append([ [], [], [], []])
                        #print(theMC,'added to all MCs',allMCs)
                    mc_idx=allMCs.index(theMC)
                    weight=1./len(thecell_MC.value.split(";"))
                    referees_for_all_MCs[mc_idx].append([ thecell_id.value , weight ])
                    referees_for_all_MCs_by_continent[mc_idx][the_continent].append([ thecell_id.value ,  weight ])

def stats_for_all_MC():
    global allMCs
    global referees_for_all_MCs
    global referees_for_all_MCs_by_continent
    global n_referees
    global n_student_referees
    
    assign_referees_to_MC()
    
    for iMC in range(len(allMCs)):        
        print('MC :',allMCs[iMC])
        weighted_all=np.sum([ theref[1] for theref in referees_for_all_MCs[iMC] ])
        weighted_eu=np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][0] ])
        weighted_as=np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][1] ])
        weighted_am=np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][2] ])
        print('Referees ',len(referees_for_all_MCs[iMC]),weighted_all,'Europe (EMEA) ',len(referees_for_all_MCs_by_continent[iMC][0]),weighted_eu,'Asia ', len(referees_for_all_MCs_by_continent[iMC][1]),weighted_as,'Americas ', len(referees_for_all_MCs_by_continent[iMC][2]),weighted_am)
    print("---")
    for theMC in range(1,8):
        if theMC == 0:
            theMCname="None"
        else:
            theMCname="MC"+str(theMC)        
        #print(theMCname)
        iMC=allMCs.index(theMCname)
        #print(iMC)
        weighted_all=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs[iMC] ]),2)
        weighted_eu=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][0] ]),2)
        weighted_as=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][1] ]),2)
        weighted_am=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][2] ]),2)
        print(theMCname,'Referees ',len(referees_for_all_MCs[iMC]),weighted_all,'Europe (EMEA) ',len(referees_for_all_MCs_by_continent[iMC][0]),weighted_eu,'Asia ', len(referees_for_all_MCs_by_continent[iMC][1]),weighted_as,'Americas ', len(referees_for_all_MCs_by_continent[iMC][2]),weighted_am)
    for theMC8 in range(1,12):
        theMCname="MC8.U"+str(theMC8).zfill(2)        
        #print(theMCname)
        if theMCname in allMCs:
            iMC=allMCs.index(theMCname)
            #print(iMC)
            weighted_all=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs[iMC] ]),2)
            weighted_eu=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][0] ]),2)
            weighted_as=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][1] ]),2)
            weighted_am=np.round(np.sum([ theref[1] for theref in referees_for_all_MCs_by_continent[iMC][2] ]),2)
            print(theMCname,'Referees ',len(referees_for_all_MCs[iMC]),weighted_all,'Europe (EMEA) ',len(referees_for_all_MCs_by_continent[iMC][0]),weighted_eu,'Asia ', len(referees_for_all_MCs_by_continent[iMC][1]),weighted_as,'Americas ', len(referees_for_all_MCs_by_continent[iMC][2]),weighted_am)
        else:
            print("No referee for ", theMCname)
    print('Number of referees',n_referees)
    print('Number of student referees (not included above)',n_student_referees)


def randomly_suggest_two_referees(MCtxt,continent,paper_id,n_referees_needed=2):
    global allMCs
    global referees_for_all_MCs_by_continent
    
    if (len(allMCs)<2):
        assign_referees_to_MC()        
    if "MC6" in MCtxt or "MC7" in MCtxt or "MC8" in MCtxt:
        MCtxt=MCtxt[0:7]
    else:
        MCtxt=MCtxt[0:3]
    #print('MCtxt',MCtxt)
    #print(allMCs)
    if MCtxt in allMCs:
        iMC=allMCs.index(MCtxt)
    else:
        iMC=-1
    #print('iMC',iMC)
    #print(allMCs[iMC])
    avalaible_referees=[]
    if iMC==-1:
        avalaible_referees=[]
    elif len(continent)==1:
        if continent[0]==0:
            avalaible_referees = referees_for_all_MCs_by_continent[iMC][1] + referees_for_all_MCs_by_continent[iMC][2]
        else:
            avalaible_referees = referees_for_all_MCs_by_continent[iMC][0]
    elif len(continent)==2:
        avalaible_referees=[]
        for icont in range(0,3):
            if icont not in continent:
                avalaible_referees=avalaible_referees+referees_for_all_MCs_by_continent[iMC][icont]
    else:
        print("!!!Warning paper with authors in 3 regions!!!")
        #exit()
        for icont in range(0,3):
            avalaible_referees=avalaible_referees+referees_for_all_MCs_by_continent[iMC][icont]

    [n_ref,the_refs]=check_referees_for_paper(paper_id)
    for the_ref in the_refs:
        print("This paper ",paper_id," has already",the_ref)
        if str(the_ref[0]) in [ str(ref[0]) for ref in avalaible_referees]:
            idx=([ str(the_ref[0]) for ref in avalaible_referees]).index(str(the_ref[0]))
            avalaible_referees.remove(avalaible_referees[idx])
            print("removing ref",the_ref)
        
    [n_declined,list_declined]=check_paper_for_declined(paper_id)
    for decl in list_declined:
        print("This paper ",paper_id," was declined by",decl)
        #print('avalaible_referees',avalaible_referees)
        print([ ref[0] for ref in avalaible_referees])
        if str(decl) in [ str(ref[0]) for ref in avalaible_referees]:
            idx=([ str(ref[0]) for ref in avalaible_referees]).index(str(decl))
            avalaible_referees.remove(avalaible_referees[idx])
            #print('avalaible_referees',avalaible_referees)
        else:
            print("not in list")
        #print('avalaible_referees',avalaible_referees)
    print('There are ', len(avalaible_referees), ' referees available for this paper ')
    if (len(avalaible_referees)<n_referees_needed):
        print("Not enough referees available for this contribution")
        print("avalaible_referees",avalaible_referees)
        selected_refs=[iref[0] for iref in avalaible_referees]
    else:
        list_refs=[ val[0] for val in avalaible_referees ] 
        weight_refs=[ val[1] for val in avalaible_referees ] 
        #print('list_refs', list_refs)
        #print('weight_refs',weight_refs)
        selected_refs=random.choices(list_refs,weights=weight_refs,k=n_referees_needed)
        if len(selected_refs)==2 and selected_refs[0]==selected_refs[1]:
            print(selected_refs)
            print("Two referees are identical... Trying again")
            print("avalaible_referees",len(avalaible_referees),avalaible_referees)
            if len(avalaible_referees)==n_referees_needed:
                selected_refs=[iref[0] for iref in avalaible_referees]
            else:
                selected_refs=random.choices(list_refs,weights=weight_refs,k=n_referees_needed)
            print(selected_refs)
            if selected_refs[0]==selected_refs[1]:
                print("Two referees are identical... Keeping only one...")
                selected_refs=[selected_refs[0]]
        print('selected_refs',selected_refs)
        #for the_ref in selected_refs:
        #    get_referee_by_id(the_ref)

        #remove assigned referees from avauilable referee

    for the_ref in selected_refs:
        #print('looking for ',the_ref)
        for icont in range(0,3):
            for iMC in range(len(allMCs)):        
                #print(referees_for_all_MCs_by_continent[iMC][icont])
                #print([ ref[0] for ref in referees_for_all_MCs_by_continent[iMC][icont]])
                if the_ref in [ ref[0] for ref in referees_for_all_MCs_by_continent[iMC][icont]]:
                    #print("found")
                    idx=([ ref[0] for ref in referees_for_all_MCs_by_continent[iMC][icont]]).index(the_ref)
                    #print(idx)
                    #print(referees_for_all_MCs_by_continent[iMC][icont])
                    referees_for_all_MCs_by_continent[iMC][icont].pop(idx)
                    #print(referees_for_all_MCs_by_continent[iMC][icont])
    return selected_refs

def load_contribs(evtid=None):
    global data_json_contribs
    if evtid==None:
        evtid=event_id
    data_json_contribs=None 
    fname=f'data/all_contribs_{evtid}.json'
    if os.path.isfile(fname):
        file_age=time.time()-os.path.getmtime(fname)
    else:
        print("Contribution file did not exist...")
        file_age=999999
    #print('file_age',file_age)
    if file_age > (24*3600):
        print('file_age',file_age)
        use_online=True
        print("Updating contribs from online data.")
    else:
        use_online=False

    if use_online:
        headers = {'Authorization': f'Bearer {api_token}'}
        data_url= f'https://indico.jacow.org/export/event/{evtid}.json?detail=contributions&pretty=yes'
        print('Data URL=',data_url)
        try:
            data = requests.get(data_url, headers=headers)
        except:
            print("Unable to access referees file")
            exit()
            
        if not data.status_code == 200:
            print("Error trying to access the conttibution data")
            print("URL was: ",data_url)
            print("Status code:",data.status_code)
            if data.text.find("invalid_token")>0:
                print("Invalid access key")
            elif data.text.find("error-box")>0:
                error_idx=data.text.find("error-box")
                print(data.text[error_idx:error_idx+500])
            exit()
        data_json_these_contribs=data.json()
        print('get done', flush=True)
        
        save_data=True
        #save_data=False
        if save_data:
            fdata=open(fname,"w")
            json.dump(data_json_these_contribs,fdata)
            fdata.close()
    else:
        fdata=open(fname,"r")
        data_json_these_contribs=json.load(fdata)
        fdata.close()
    if evtid==event_id:
        data_json_contribs=data_json_these_contribs
    return data_json_these_contribs

def find_contrib(the_id=None,the_db_id=None):
    global data_json_contribs
    load_contribs()
    
    result_json=data_json_contribs["results"][0]
    for contrib in result_json['contributions']:
        if ((the_id is not None and contrib['id'] == str(the_id)) or (the_db_id is not None and contrib['db_id']== int(the_db_id))):
            #print(the_id)
            #print(contrib['id'])
            #print(the_db_id)
            #print(contrib['db_id'])
            return contrib
    #print("Not found id=", the_id," db_id=", the_db_id)
    return None

def remind_referees():
    global all_actions_todo
    days_after_initial_message=2
    days_between_reminders=1
    days_before_deadline_reminder=3

    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()
    
    #Looks for the line where to enter the data
    irow=2
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        #print(str(referees_assignation_wb.cell(row=irow,column=6).value))
        date_status=datetime.datetime.now()-datetime.datetime.strptime(str(referees_assignation_wb.cell(row=irow,column=7).value),"%Y-%m-%d %H:%M:%S")
        referee_id=str(referees_assignation_wb.cell(row=irow,column=3).value)
        ref_name=str(referees_assignation_wb.cell(row=irow,column=4).value)
        #print('date_status',date_status.days,ref_name)
        contrib_id=str(referees_assignation_wb.cell(row=irow,column=1).value)        
        contrib=find_contrib(the_id=contrib_id)
        if contrib is None:
            print("contrib ",contrib_id," is None but has a referee assigned!")
            print("Status",str(referees_assignation_wb.cell(row=irow,column=6).value))
            if not (str(referees_assignation_wb.cell(row=irow,column=6).value) == "Withdrawn") and not (str(referees_assignation_wb.cell(row=irow,column=6).value) == "Declined") and not (str(referees_assignation_wb.cell(row=irow,column=6).value) == "Review received"):
                exit()
        if str(referees_assignation_wb.cell(row=irow,column=6).value)=="Email request sent":
            no_reply=True
            if referees_assignation_wb.cell(row=irow,column=9).value is not None:
                if len(referees_assignation_wb.cell(row=irow,column=9).value)>3:
                    no_reply=False
            if referees_assignation_wb.cell(row=irow,column=10).value is not None:
                if len(referees_assignation_wb.cell(row=irow,column=10).value)>3:
                    no_reply=False
            if no_reply:
                date_request=datetime.datetime.now()-datetime.datetime.strptime(str(referees_assignation_wb.cell(row=irow,column=7).value),"%Y-%m-%d %H:%M:%S")
                #print(date_request.days)
                if date_request.days>days_after_initial_message:
                    print("Acceptance overdue",contrib_id)
                    entry_code=str(((int(referee_id)*10000)+int(contrib_id))*7)
                    #print(entry_code)
                    urlassign="http://nicolas.delerue.org/ipac23/assign.php?"
                    print('name',ref_name)
                    print('title',contrib['title'])
                    urlassign=urlassign+urllib.parse.urlencode({'key': str(entry_code), 'title': contrib['title'], 'name' : ref_name })
                    #print('url',urlassign)
                    data = requests.get(urlassign)
                    if not data.status_code == 200:
                        print("Error submitting the assignememnt: wrong response code")
                    else:
                        url="http://nicolas.delerue.org/ipac23/referee_acceptance_form.php?id="+entry_code
                        ef.send_email(referee_id=referee_id,paper_id=contrib_id,msgfile='message_referee_reminder.txt',url=url)
                        referees_assignation_wb.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")        
                        referees_assignation_wb.cell(row=irow,column=6).value="Reminder sent"
                        wb_obj.save(file)
                    #time.sleep(1)
        elif (str(referees_assignation_wb.cell(row=irow,column=6).value)=="Reminder sent"):
            #print('Reminder sent')
            if date_status.days>days_between_reminders:
                print("Second reminder to be sent to ",referee_id,ref_name)
                entry_code=str(((int(referee_id)*10000)+int(contrib_id))*7)
                print('name',ref_name)
                print('title',contrib['title'])
                url="http://nicolas.delerue.org/ipac23/referee_acceptance_form.php?id="+entry_code
                ef.send_email(referee_id=referee_id,paper_id=contrib_id,msgfile='message_referee_second_reminder.txt',url=url)
                referees_assignation_wb.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                referees_assignation_wb.cell(row=irow,column=6).value="2nd Reminder sent"
                wb_obj.save(file)
                #time.sleep(1)
        elif (str(referees_assignation_wb.cell(row=irow,column=6).value)=="2nd Reminder sent"):
            if date_status.days>days_between_reminders:
                print("Paper to be reassigned: ",contrib_id," ref ",referee_id,ref_name)
                print("./assign_referee.py --paper ",contrib_id," --referee",referee_id," --overdue")
                all_actions_todo.append(["Paper to be reassigned: "+str(contrib_id)+" ref: "+str(referee_id)+" "+str(ref_name) , "./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(referee_id)+"  --overdue"] )
                #print(contrib)
                print(contrib['track'])
                print("~~~")
                #time.sleep(10)
                #exit()
        elif (str(referees_assignation_wb.cell(row=irow,column=6).value)=="Accepted"):    
            if date_status.days>days_between_reminders:
                deadline_cell=str(referees_assignation_wb.cell(row=irow,column=8).value)
                if not "00:00" in deadline_cell: 
                    date_deadline=datetime.datetime.now()-datetime.datetime.strptime(deadline_cell,"%d/%m/%Y")
                else:
                    date_deadline=datetime.datetime.now()-datetime.datetime.strptime(deadline_cell,"%Y-%m-%d %H:%M:%S")         
                #print('date_deadline', date_deadline.days)
                if date_deadline.days-days_before_deadline_reminder>0:
                    print("Deadline coming soon for ",referee_id,ref_name)
                    entry_code=str(((int(referee_id)*10000)+int(contrib_id))*7)
                    url="http://nicolas.delerue.org/ipac23/referee_acceptance_form.php?id="+entry_code
                    ef.send_email(referee_id=referee_id,paper_id=contrib_id,msgfile='message_referee_deadline_soon.txt',url=url)
                    referees_assignation_wb.cell(row=irow,column=7).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")        
                    referees_assignation_wb.cell(row=irow,column=6).value="Deadline reminder sent"
                    wb_obj.save(file)
                    time.sleep(3)
        irow=irow+1
            
def get_referees_list():            
    headers = {'Authorization': f'Bearer {api_token}'}
    data = requests.get(f'{event_url}manage/papers/teams/', headers=headers)
    for line in data.text.split("\\n"):
        if 'id=\\"content_reviewers' in line:
            return line.replace("&#34;","").replace("User:","").split("[")[1].split("]")[0].split(",")
    return None
            
#### End referees

#### Countries
def get_countries_from_contrib(contrib):
    continent=[]
    countries=[]
    affiliations=[]
    #print('---')
    for type in [ 'speakers' , 'primaryauthors' , 'coauthors' ]: 
        for speak in contrib[type]:
            affiliations.append(speak['affiliation'])
            #if type == 'speakers':
                #print(type,speak['first_name'],speak['last_name'])
    for affiliation in affiliations:
        country=get_country_from_affiliation(affiliation)
        if country is None or len(country.strip())<2:
            print("No country for affiliation",affiliation)
            find_new_affilation(contrib['db_id'])
            country=get_country_from_affiliation(affiliation)
        if country is None:
            print('Country not found despit trying for affiliation', affiliation)
            load_affiliations_countries_file()
        else:
            countries.append(country)

    countries=list(set(countries))
    for country in countries:
        #print(country,get_continent(country))
        continent.append(get_continent(country))

    retdict={}
    retdict['countries']=countries
    retdict['continent']=continent
    return retdict
            
all_countries=[]
def get_continent(country):
    global all_countries

    #print('get continent')
    europe_countries=['France', 'United Kingdom', 'Switzerland', 'Germany', 'Italy', 'Sweden' , 'Poland' , 'Ghana' ,  'Spain', 'Czechia', 'Romania' , 'Austria', 'Belgium', 'Russia' , 'Israel' , 'Turkey' , 'The Netherlands' , 'Netherlands', 'Latvia' , 'Croatia' ,'Palestine' , 'Hungary' , 'Jordan' , 'South Africa' , 'Portugal' , 'Slovenia' , 'Greece' ]
    asia_countries=['Iran', 'China', 'Taiwan', 'Thailand', 'Japan', 'India' , 'South Korea' , 'Armenia' , 'Australia' ]
    america_countries=[ 'United States' , 'Canada' , 'Brazil' ]
    if country in europe_countries:
        return 0
    elif country in asia_countries:
        return 1
    elif country in america_countries:
        return 2
    elif country is None:
        return 3
    else:
        print('Country ',country,' is not in any continent')
        all_countries.append(country)
        print(all_countries)
        exit()
affiliation_countries=[]

def load_affiliations_countries_file():
    global affiliation_countries
    if len(affiliation_countries)==0:
        fname="affiliations_countries.txt"
        fdata=open(fname,"r")
        lines=fdata.readlines()
        fdata.close()
        for line in lines:
            #print(line)
            if len(line)>3:
                vals=line.strip().split(";")
                affiliation_countries.append(vals)
        #print(affiliation_countries)

no_country_affiliation=[]
def get_country_from_affiliation(affiliation):
    global affiliation_countries
    global no_country_affiliation
    #print("Looking for country for affiliation", affiliation)
    load_affiliations_countries_file()
    for aff in affiliation_countries:
        #print(aff[0].strip(),affiliation.strip())
        if aff[0].lower().strip() == affiliation.lower().strip():
            #print("Match",aff[0].strip(),aff[1])
            return aff[1].strip()
    if len(affiliation.strip())>2:
        print("No match for affiliation ", affiliation)
        no_country_affiliation.append(affiliation)
        #exit()
    return None

def find_new_affilation(contrib_db_id):
    data_json=get_paper_info(contrib_db_id)
    if data_json is not None:
        print("New affiliation found")
        try:
            aff=data_json['revisions'][0]['submitter']['affiliation_meta']['name']
            country=data_json['revisions'][0]['submitter']['affiliation_meta']['country_name']
            print(aff,';',country)
            fname="affiliations_countries.txt"
            fdata=open(fname,"a")
            fdata.write(aff+";"+country+"\n")
            fdata.close()
        except:
            print("Unable to identify affiliation")
        load_affiliations_countries_file()

### End countries

### Papers and contribs 
def get_paper_info(db_id,use_cache=False,sleep_before_online=0.5):
    filename="papers/paper_"+str(db_id)+".json"
    use_online=True
    if use_cache:
        #print(filename)
        if os.path.isfile(filename):
            file_age=time.time()-os.path.getmtime(filename)
            if file_age > (24*60*60):
                print('file_age',file_age)
                use_online=True
                print("Updating paper from online")
            else:
                use_online=False
        else:
            use_online=True
    #print('use_online',use_online)
    if use_online:
        time.sleep(sleep_before_online)
        headers = {'Authorization': f'Bearer {api_token}'}
        url=f'{event_url}papers/api/'+str(db_id)
        data = requests.get(url, headers=headers)
        #print('accessing url ',url)
        if data.status_code == 200:
            fdata=open(filename,"w")
            json.dump(data.json(),fdata)
            fdata.close()
            return(data.json())
        else:
            fdata=open(filename,"w")
            fdata.write("{}")
            fdata.close()
            print('Status code (paper access)', data.status_code)
            #print('db_id',db_id)
            return None
    else:
        fdata=open(filename,"r")
        data_json=json.load(fdata)
        fdata.close()
        #print('data_json',data_json,len(data_json))
        if len(data_json) <5:
            return None
        else:
            return data_json
        
n_reviews=[]
contribs_accepted=[]
contribs_rejected=[]
submitted_contribs_list=[]
to_be_corrected_contribs=[]
cpt_reviews=[ -1, -1, -1 ]

def submitted_contribs(evtid=None,force_online=False):
    global n_reviews
    global contribs_accepted
    global contribs_rejected
    global submitted_contribs_list
    global to_be_corrected_contribs
    #global cpt_reviews
    submitted_contribs_list=[]
    to_be_corrected_contribs=[]
    contribs_accepted=[]
    contribs_rejected=[]
    
    if evtid==None:
        evtid=event_id
    fname=f'data/submitted_contribs_{evtid}.html'
    if os.path.isfile(fname):
        file_age=time.time()-os.path.getmtime(fname)
    else:
        file_age=999999
    if file_age > (1*60*60) or force_online :
        print('file_age',file_age)
        use_online=True
        print("Updating submitted contribs from online")
        verify_contrib_consistency=True
    else:
        use_online=False
        verify_contrib_consistency=False

        
    if use_online:
        #Reads a submitted contribution based on its ID
        headers = {'Authorization': f'Bearer {api_token}'}
        data_url=f'https://indico.jacow.org/event/{evtid}/manage/papers/assignment-list/'
        data = requests.get(data_url, headers=headers)
        if data.status_code == 200:
            fdata=open(fname,"w")
            fdata.write(data.text)
            fdata.close()
            data_text=data.text
        else:
            print("Error when reading ",data_url)
            print("Error code",data.status_code)
            data_text=""
    else:
        fdata=open(fname,"r")
        data_text=fdata.read()
        fdata.close()

    
    all_lines=data_text.split("\n")
    contrib_id=None
    unsubmitted_contribs=[]
    n_reviews=[]
    #cpt_reviews=[ 0, 0, 0 ]
    n_accepted=0
    n_rejected=0
    search_review=False
    search_revision=False
    search_judge=False
    search_reviewer=False
    for line in all_lines:
        #print('line',line)
        if "contrib-" in line:
            if contrib_id is not None:
                print('New contrib before status of the previous one; check status of contrib',contrib_id)
            #print('contrib',line)
            contrib_id=line.split('"')[1].split("-")[1]
            #print('contrib_id',contrib_id)
        if "data-searchable"in line and ("ubmitted" in line or "to be corrected" in line or "accepted" in line or "rejected" in line ) and not "data-text" in line:
            if contrib_id is None:
                print('submitted',line)
                print("Status before contrib!")
                exit()
            if '"paper not yet submitted"' in line:
                unsubmitted_contribs.append(contrib_id)
                contrib_id=None
            elif 'Not submitted' in line:
                unsubmitted_contribs.append(contrib_id)
                contrib_id=None
            elif '"submitted"' in line:
                submitted_contribs_list.append(contrib_id)
                n_reviews.append(-1)
                contrib_id=None
                search_review=True
            elif '"to be corrected"' in line:
                to_be_corrected_contribs.append(contrib_id)
                contrib_id=None
            elif '"accepted"' in line:
                #print("Paper accepted")
                n_accepted=n_accepted+1
                contrib_id=None
                contribs_accepted.append(contrib_id)
            elif '"rejected"' in line:
                #print("Paper rejected")
                contrib_id=None
                n_rejected=n_rejected+1
                contribs_rejected.append(contrib_id)
            else:
                print("Not understood",line)
                exit()
        if search_review:
            if "0 reviews" in line:
                n_reviews[-1]=0
                search_review=False
                search_revision=True
            elif "1 review" in line:
                n_reviews[-1]=1
                search_review=False
                search_revision=True
            elif "2 reviews" in line:
                n_reviews[-1]=2
                search_review=False
                search_revision=True
            elif "3 reviews" in line:
                n_reviews[-1]=3
                search_review=False
                search_revision=True
            if search_review==False:
                if n_reviews[-1]>0:
                    if verify_contrib_consistency:
                        check_reviews_for_contrib(submitted_contribs_list[-1],n_reviews[-1])
                #print('search_revision',search_revision)
        if search_revision and "revision-column" in line:
            search_revision=False
            search_judge=True
            search_reviewer=False
        if search_reviewer and "data-searchable" in line:
            search_judge=False
            search_reviewer=False
            the_reviewer=line.split('"')[1].strip()
            #print('the_reviewer',the_reviewer)
            #if verify_contrib_consistency:
            #    check_contrib_consistency(submitted_contribs[-1],the_judges,the_reviewer)
        if search_judge and "data-searchable" in line:
            search_judge=False
            search_reviewer=True
            the_judges=line.split('"')[1].strip()
            #print('the_judges',the_judges)

    print('Submitted contribs: ',len(submitted_contribs_list))
    print('Unsubmitted contribs: ',len(unsubmitted_contribs))
    print('To be corrected contribs', len(to_be_corrected_contribs))
    print('Accepted ', n_accepted)
    print('Rejected ', n_rejected)

    return submitted_contribs_list

def print_contrib_stats():
    global n_reviews
    global contribs_accepted
    global contribs_rejected
    global submitted_contribs_list
    global to_be_corrected_contribs
    global cpt_reviews

    print("### Statistics ###")
    print("* Stats from contrib file *")
    print('len(submitted_contribs_list)',len(submitted_contribs_list))
    #print('n_reviews',n_reviews)
    print('cpt_reviews',cpt_reviews)
    print('len(contribs_accepted)',len(contribs_accepted))
    print('len(contribs_rejected)',len(contribs_rejected))

    print('len(to_be_corrected_contribs)',len(to_be_corrected_contribs))
    
    
def check_all_papers():
    global cpt_reviews
    global all_actions_todo
    global n_days_to_resubmit_after_1st_round
    submitted_papers=[]
    unsubmitted_papers=[]
    state_submitted_papers=[]
    accepted_papers=[]
    to_be_corrected_papers=[]
    second_round_papers=[]
    rejected_papers=[]
    cpt_reviews=[0, 0, 0, 0, 0, 0]
    cpt_referees=[0, 0, 0, 0, 0, 0, 0]
    n_second_round=[0 , 0, 0 , 0 , 0, 0]
    two_reviews_on_second_round=[]
    papers_with_less_than_two_refs=[]
    not_none=0
    paper_number=500
    fname="check_all_papers.id"
    #if os.path.isfile(fname):
    #    fdata=open(fname,"r")
    #    paper_number=int(fdata.readlines()[0])
    #    fdata.close()
    #print('paper_number',paper_number)
    for the_id in range(paper_number,2800):
        #for the_id in range(2020,2050):
        fdata=open(fname,"w")
        fdata.write(str(the_id)+"\n")
        fdata.close()
        contrib=find_contrib(the_id)
        if contrib is not None:
            #print('Checking paper for contrib ', the_id,' that is not none; db_id ',contrib['db_id'])
            not_none=not_none+1
            the_paper=get_paper_info(contrib['db_id'],use_cache=True,sleep_before_online=0.1)
            if the_paper is None:
                #print("not submitted")
                unsubmitted_papers.append(the_id)
            else:
                print('Contribution submitted: ', the_id,' that is not none; db_id ',contrib['db_id'])
                #print('len(the_paper)',len(the_paper))
                if len(the_paper)<3:
                    print("short paper")
                    exit()
                in_second_round=check_paper_in_double_reviews(the_id)
                
                if in_second_round:
                    secound_round_data=get_paper_in_double_reviews(the_id)
                    if secound_round_data['suggestion_spb_2nd_round'] is not None and len(secound_round_data['suggestion_spb_2nd_round'])>5:
                        if secound_round_data['date_authors_notified_2nd_round'] is not None and (len(secound_round_data['date_authors_notified_2nd_round'])>5):
                            print("Review process complete",secound_round_data['decision_spb_2nd_round'])
                        else:
                            #print(len(secound_round_data['suggestion_spb_2nd_round']))
                            #print("A decision has already been suggested for this paper on ",str(secound_round_data['date_suggestion_spb_2nd_round']))
                            age_notif=datetime.datetime.now()-datetime.datetime.strptime(str(secound_round_data['date_suggestion_spb_2nd_round']),"%Y-%m-%d %H:%M:%S")
                            #print("SPB was notified ",age_notif.days," ago",secound_round_data['suggestion_spb_2nd_round'])
                            if age_notif.days>=2:
                                all_actions_todo.append(["Secound round decided", " ./find_paper.py --id "+str(the_id)+"  --notify-authors" ])
                    else:
                        print("Paper is in secound round")
                        n_second_round[0]=n_second_round[0]+1
                        if int(the_paper['last_revision']['number'])>int(secound_round_data['revision']):
                            print("A new version of paper ", the_id," has been submitted")
                            n_second_round[1]=n_second_round[1]+1
                            second_round_papers.append(the_id)
                            check_resubmitted_paper(the_paper)
                        elif secound_round_data['decision_1st_round']  == "To be corrected" or secound_round_data['decision_1st_round'] == "Minor revisions (no second round)":
                            age_notif=datetime.datetime.now()-datetime.datetime.strptime(str(secound_round_data['date_notification']),"%d/%m/%Y %H:%M")
                            print("author was notified ",age_notif.days," ago and has not resubmitted")
                            if age_notif.days>n_days_to_resubmit_after_1st_round and (secound_round_data['date_author_resubmission_reminder'] is None or len(secound_round_data['date_author_resubmission_reminder'])<5):
                                new_data={}
                                new_data['date_author_resubmission_reminder']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                
                                substitution_dict={}
                                substitution_dict['title']=contrib['title']
                                substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
                                substitution_dict['url_paper']="{event_url}papers/"+str(the_paper['contribution']['id'])+"/"
                    
                                substitution_dict['submitter']=str(the_paper['last_revision']['submitter']['full_name'])+' '+str(the_paper['last_revision']['submitter']['email'])
                    
                                substitution_dict['reviewers_comments']=paper_review_report_txt(the_paper,for_spb=False,revision_number=secound_round_data['revision'])
                                
                                ef.email_file(the_paper['last_revision']['submitter']['email'],'message_author_paper_not_yet_resubmitted.txt',replace_dict=substitution_dict)
                                write_paper_in_double_reviews(the_paper['contribution']['friendly_id'],new_data)
                                #exit()
                else:
                    secound_round_data=None
                #print("submitted")
                submitted_papers.append(the_id)
                if the_paper['state']['name'] == 'submitted':
                    state_submitted_papers.append(the_id)
                    refs=check_referees_for_paper(the_id,show_row=False)
                    n_refs=int(refs[0])
                    cpt_referees[n_refs]=cpt_referees[n_refs]+1
                    if n_refs <2:
                        papers_with_less_than_two_refs.append(the_id)
                elif the_paper['state']['name'] == 'accepted':
                    accepted_papers.append(the_id)
                elif the_paper['state']['name'] == 'to_be_corrected':
                    to_be_corrected_papers.append(the_id)
                elif the_paper['state']['name'] == 'rejected':
                    rejected_papers.append(the_id)
                else:
                    print("Unknown paper state:", the_paper['state']['name'])
                    exit()

                n_reviews=0
                n_comments=0                
                for timel in the_paper['last_revision']['timeline']:
                    if timel['timeline_item_type']=="comment":
                        #print(timel['text'])
                        n_comments=n_comments+1
                    elif timel['timeline_item_type']=="review":
                        n_reviews=n_reviews+1
                        the_ref=find_referee_by_email(timel['user']['email'])
                        print("Review by ", timel['user']['email'], the_ref[0])
                        if len(the_ref)==1:
                            #print("checking file", the_id,the_ref[0])
                            the_line=referee_assignation_file_get_line(the_id,the_ref[0])
                            #print(the_line)
                            if not (the_line['status'] in [ 'Review received' , 'Withdrawn' ] ):
                                print("Review was not recorded for paper ", the_id,"  ref ",the_ref[0]) 
                                #print(the_line)
                                #print(" ./find_paper.py --id ",the_id," --record-review")
                                notification_review_received(the_ref[0],the_id,timel['proposed_action']['name'],timel['created_dt'],the_paper['contribution']['title'])
                                #exit()
                            else:
                                #check that there no second line for this entry
                                if referee_assignation_file_get_line(the_id,the_ref[0],int(the_line['row'])+1):
                                    print("Alert there seem to be more than one line paper ", the_id,"  ref ",the_ref[0])
                        else:
                            print('the_ref does not have length 1',the_ref)
                            exit()
                cpt_reviews[n_reviews]=cpt_reviews[n_reviews]+1
                if in_second_round:
                    if int(the_paper['last_revision']['number'])>int(secound_round_data['revision']):
                        n_second_round[2+n_reviews]=n_second_round[2+n_reviews]+1
                if n_reviews>=2:
                    if not check_paper_in_double_reviews(the_id):
                        print("Paper ", the_id , "is not in double review file")
                        print(" ./find_paper.py --id ",the_id)
                        all_actions_todo.append(" ./find_paper.py --id "+str(the_id))
                    if in_second_round:
                        if int(the_paper['last_revision']['number'])>int(secound_round_data['revision']):
                            print("Paper ", the_id , " has two reviews on second round ")
                            two_reviews_on_second_round.append(the_id)
                            #time.sleep(1)
                #### Status update
                #Set to True to send a status update to all submitted papers
                if False:
                    already_notified=False
                    for timel in the_paper['last_revision']['timeline']:
                        print("Looking at timeline", timel['timeline_item_type'])
                        if timel['timeline_item_type']=="comment":
                            if timel['user']['email'] == "delerue@lal.in2p3.fr" and "Review update on your paper" in timel['text']:
                                print("Already notified")
                                #time.sleep(10)
                                already_notified=True
                    if the_paper['state']['name'] == 'submitted' and not already_notified and not check_paper_in_double_reviews(the_id):
                        author_email=the_paper['last_revision']['submitter']['email']
                        msg_txt="Review update on your paper '"+the_paper['contribution']['title']+"'\n"
                        msg_txt=msg_txt+"Dear "+the_paper['last_revision']['submitter']['full_name']+",\n\n"
                        msg_txt=msg_txt+'Thank you for submitting your paper "'+the_paper['contribution']['title']+'" to the IPAC\'23 Light Peer Review. To date the review of your paper is not complete.\n'
                        [n_declined,list_declined]=check_paper_for_declined(the_id)
                        if n_declined>0:
                            if n_declined==1:
                                msg_txt=msg_txt+"We have submitted your paper for review but one of the reviewer declined reviewing it so another reviewer has recently been appointed.\n"
                            else:
                                msg_txt=msg_txt+"We have submitted your paper for review but "+str(n_declined) +" reviewers declined reviewing it so other reviewers had to be appointed.\n"
                        if n_reviews == 0:
                            msg_txt=msg_txt+"We have not yet receievd any reviews.\n"
                        elif n_reviews == 1:
                            msg_txt=msg_txt+"We have already received one review and are waiting for the second one.\n"
                        else:
                            if not check_paper_in_double_reviews(the_id):
                                msg_txt=msg_txt+"We have already received two reviews but had to seek the opinion of a third reviewer which is awaited.\n"
                            else:
                                msg_txt=msg_txt+"We have already received two reviews and these review are under consideration but the Scientific Publication Board.\n"                               
                        msg_txt=msg_txt+"\nWe will keep you informed once the Scientific Publication Board has sufficient reviews to take a decision on your paper.\n"
                        msg_txt=msg_txt+"\nBest wishes,\n"
                        msg_txt=msg_txt+"Nicolas Delerue, on behalf of the Scientific Publication Board of IPAC'23\n\n"
                        msg_txt=msg_txt+"Contribution id: "+str(the_id)
                        print(msg_txt)
                        comment_paper(contrib['db_id'],msg_txt)
                        ef.email_txt(author_email,msg_txt)
                        the_paper=get_paper_info(contrib['db_id'],use_cache=False,sleep_before_online=2)
                        #time.sleep(10)
    fdata=open(fname,"w")
    fdata.write(str(500)+"\n")
    fdata.close()

    
    print('papers_with_less_than_two_refs',papers_with_less_than_two_refs)
    print('two_reviews_on_second_round',two_reviews_on_second_round)
    print("----")
    print('Submitted papers: ',len(submitted_papers))
    print('Unsubmitted papers: ',len(unsubmitted_papers))
    print('To be corrected contribs', len(to_be_corrected_papers))
    print('n_second_round',n_second_round)
    print('len(two_reviews_on_second_round)',len(two_reviews_on_second_round))
    print('State: submitted',len(state_submitted_papers))
    print('Referees counter',cpt_referees)
    print('papers_with_less_than_two_refs',len(papers_with_less_than_two_refs))
    print('Review counter', cpt_reviews)
    print('Accepted',len(accepted_papers))    
    print('to_be_corrected papers',len(to_be_corrected_papers))
    print('Rejected',len(rejected_papers))

    #if len(all_actions_todo)>0:
    #print("*** All actions to do ***")
    #print(all_actions_todo)
    #exit()

def get_substitution_dict_for_paper(contrib):
    the_paper=get_paper_info(contrib['db_id'],use_cache=True,sleep_before_online=2)
    substitution_dict={}
    substitution_dict['title']=contrib['title']
    substitution_dict['paper_id']=the_paper['contribution']['friendly_id']
    substitution_dict['url_paper']="{event_url}papers/"+str(the_paper['contribution']['id'])+"/"
        
    substitution_dict['submitter']=str(the_paper['last_revision']['submitter']['full_name'])+' '+str(the_paper['last_revision']['submitter']['email'])
        
    #substitution_dict['reviewers_comments']=paper_review_report_txt(the_paper,for_spb=False,revision_number=secound_round_data['revision'])
    return substitution_dict

    
STATUS_COL=6
REVIEW_STATUS_COL=10
REVIEW_ACTION_COL=11
REVIEW_DATE_COL=12
REF_ID_COL=3
REF_NAME_COL=4
PAPER_COL=1
def check_reviews_for_contrib(contrib_db_id,n_reviews):
    global REVIEW_STATUS_COL
    global REVIEW_ACTION_COL
    global REVIEW_DATE_COL
    global REF_ID_COL
    global PAPER_COL

    global all_actions_todo
    
    print('contrib_db_id',contrib_db_id,'n_reviews',n_reviews)
    contrib=find_contrib(the_db_id=contrib_db_id)
    contrib_id=contrib['id']
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    referees_assignation_wb = wb_obj.active
    cpt_review=0
    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        #print(referees_assignation_wb.cell(row=irow,column=PAPER_COL).value,str(contrib_id))
        if referees_assignation_wb.cell(row=irow,column=PAPER_COL).value == str(contrib_id):
            #print("paper match")
            if referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value is not None and referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value == "Review received" :
                #print("Review line",len(referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value),referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value)
                print('Review on line ',irow)
                cpt_review=cpt_review+1
        irow=irow+1
    if not cpt_review == n_reviews:
        print('cpt_review',cpt_review)
        print('n_reviews',n_reviews)
        print('Contrib ',contrib_id,' has new reviews')
        contrib_data=find_contrib(the_id=contrib_id)
        print('db_id',contrib_data['db_id'])
        paper=get_paper_info(contrib_data['db_id'],use_cache=False,sleep_before_online=0.5)
        #print(paper)
        #print('timeline')
        review_decisions=[]
        for timel in paper['last_revision']['timeline']:
            print('type: ',timel['timeline_item_type'])
            if timel['timeline_item_type']=="review":
                email=timel['user']['email']
                the_ref=get_referee_by_email(email)              
                print("Review by ",email)
                #print(the_ref)
                the_ref_id=the_ref[5]
                #print('the_ref_id',the_ref_id)
                #print('proposed: ',timel['proposed_action']['name'])
                review_decisions.append(timel['proposed_action']['name'])
                irow=1
                while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
                    if str(referees_assignation_wb.cell(row=irow,column=REF_ID_COL).value) == str(the_ref_id) and referees_assignation_wb.cell(row=irow,column=PAPER_COL).value == str(contrib_id):
                        if referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value is not None and len(referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value)>3:
                            #print("This review was known")
                            pass
                        else:
                            notification_review_received(the_ref_id,contrib_id,timel['proposed_action']['name'],timel['created_dt'],paper['contribution']['title'])
                    irow=irow+1
        if len(review_decisions)>=2:
            print('Contribution ', contrib_id,' has completed the review process')
            if check_paper_in_double_reviews(contrib_id):
                print("Already in double review")
            else:
                print("You can check its status at {event_url}papers/"+str(paper['contribution']['id'])+"/")
                print("Or by typing:  ./find_paper.py --id "+str(contrib_id))
                print(review_decisions)
                all_actions_todo.append(['Contribution '+str(contrib_id)+' has completed the review process' , "./find_paper.py --id "+str(contrib_id) ])

def notification_review_received(the_ref_id,contrib_id,proposed_action,date_received,paper_title):
    global REVIEW_STATUS_COL
    global REVIEW_ACTION_COL
    global REVIEW_DATE_COL
    global REF_ID_COL
    global PAPER_COL
    
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 
    referees_assignation_wb = wb_obj.active
    irow=1
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        if str(referees_assignation_wb.cell(row=irow,column=REF_ID_COL).value) == str(the_ref_id) and referees_assignation_wb.cell(row=irow,column=PAPER_COL).value == str(contrib_id):
            print("line found", irow)
            if referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value is None or len(referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value)>3:
                print("recording")
                referees_assignation_wb.cell(row=irow,column=STATUS_COL).value="Review received"
                referees_assignation_wb.cell(row=irow,column=REVIEW_STATUS_COL).value="Review received"
                referees_assignation_wb.cell(row=irow,column=REVIEW_ACTION_COL).value=proposed_action                
                referees_assignation_wb.cell(row=irow,column=REVIEW_DATE_COL).value=date_received
                contrib_id=int(referees_assignation_wb.cell(row=irow,column=1).value)
                entry_code=str(((int(the_ref_id)*10000)+int(contrib_id))*7)
                urlassign="http://nicolas.delerue.org/ipac23/assign.php?"
                urlassign=urlassign+urllib.parse.urlencode({'key': str(entry_code), 'title': paper_title, 'name' : the_ref_id })                
                data = requests.get(urlassign)
                if not data.status_code == 200:
                    print("Error submitting the assignememnt: wrong response code")
                url="http://nicolas.delerue.org/ipac23/referee_additional_review.php?id="+str(entry_code)
                ef.send_email(referee_id=int(the_ref_id),paper_id=contrib_id,msgfile='message_thank_you_review.txt',url=url)
                wb_obj.save(file)
        irow=irow+1
            
papers_double_review = "papers_double_review.xlsx"
double_review_cols_structure=[[ 'suggestion' , 6]  , [ 'prab' , 7] , [ 'decision_1st_round' , 9 ], [ 'date_notification' , 10 ] , [ 'revision' ,12 ] , [ 'date_author_resubmission_reminder' ,13 ] ,  [ 'date_resubmitted' , 14 ] , [ 'date_notified_2nd_round' , 15 ] ,  [ 'date_status_rev1' , 16 ] ,  [ 'status_rev1' , 17 ] , [ 'date_2nd_review1' , 18 ] , [ 'name_2nd_review1' , 19 ] , [ 'decision_2nd_review1' , 20 ]  ,  [ 'date_status_rev2' , 21 ] ,  [ 'status_rev2' , 22 ] , [ 'date_2nd_review2' , 23 ] , [ 'name_2nd_review2' , 24 ] , [ 'decision_2nd_review2' , 25 ]  ,  [ 'date_status_rev3' , 26 ] ,  [ 'status_rev3' , 27 ] , [ 'date_2nd_review3' , 28 ] , [ 'name_2nd_review3' , 29 ] , [ 'decision_2nd_review3' , 30 ] , [ 'suggestion_spb_2nd_round' , 31 ] , [ 'date_suggestion_spb_2nd_round' , 32 ] , [ 'decision_spb_2nd_round' , 33 ] , [ 'date_authors_notified_2nd_round' , 34 ] ]

def check_paper_in_double_reviews(contrib_id):
    return get_paper_in_double_reviews(contrib_id) is not None

def get_paper_in_double_reviews(contrib_id):
    global papers_double_review
    global double_review_cols_structure
    file = papers_double_review
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    double_review_wb = wb_obj.active
    #print('Checking contrib_id in double review',contrib_id)
    retval=None
    
    irow=1
    if not "Contrib ID" in double_review_wb.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    while (double_review_wb.cell(row=irow,column=1).value is not None) and not str(double_review_wb.cell(row=irow,column=1).value) == str(contrib_id):
        irow=irow+1
    if str(double_review_wb.cell(row=irow,column=1).value) == str(contrib_id):
        #print("Paper found in double_review file")
        retval={}
        retval['row']=irow
        for the_key in double_review_cols_structure:
            retval[the_key[0]]=str(double_review_wb.cell(row=irow,column=int(the_key[1])).value)
        return retval
    else:
        #print("Paper not found in double_review file")
        return None
    return None

def write_paper_in_double_reviews(contrib_id,paper_data):
    global papers_double_review
    global double_review_cols_structure
    file = papers_double_review
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    double_review_wb = wb_obj.active
    #print('Checking contrib_id in double review',contrib_id)
    retval=None
    
    irow=1
    if not "Contrib ID" in double_review_wb.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    while (double_review_wb.cell(row=irow,column=1).value is not None) and not str(double_review_wb.cell(row=irow,column=1).value) == str(contrib_id):
        irow=irow+1
    if str(double_review_wb.cell(row=irow,column=1).value) == str(contrib_id):
        #print("Paper found in double_review file - writing", irow)
        #print(paper_data)
        for the_key in double_review_cols_structure:
            if the_key[0] in  paper_data.keys():
                double_review_wb.cell(row=irow,column=int(the_key[1])).value=paper_data[the_key[0]]
            
        wb_obj.save(file)
        #print("saved")
        return irow
    else:
        print("Error: Paper not found in double_review file")
        exit()
    print("Error: Paper not found in double_review file")
    exit()

papers_accepted= "papers_accepted.xlsx"
papers_rejected= "papers_rejected.xlsx"
def record_accepted_paper(contrib_id):
    global papers_accepted
    file = papers_accepted
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    accepted_papers = wb_obj.active

    contrib_data=find_contrib(the_id=contrib_id)
    print('db_id',contrib_data['db_id'])
    the_paper=get_paper_info(contrib_data['db_id'])
    
    irow=1
    if not "Contrib ID" in accepted_papers.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    irow=2
    while (accepted_papers.cell(row=irow,column=1).value is not None):
        irow=irow+1
    
    accepted_papers.cell(row=irow,column=1).value=contrib_id
    accepted_papers.cell(row=irow,column=2).value=contrib_data['db_id']
    accepted_papers.cell(row=irow,column=3).value=the_paper['contribution']['title']
    accepted_papers.cell(row=irow,column=4).value=the_paper['last_revision']['submitter']['full_name']
    accepted_papers.cell(row=irow,column=5).value=the_paper['last_revision']['submitter']['email']
    accepted_papers.cell(row=irow,column=6).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb_obj.save(file)

    #Unassign referees
    refs=check_referees_for_paper(the_paper['contribution']['friendly_id'])
    for the_ref in refs[1]:        
        unassign_referee(contrib_data['db_id'],get_referee_db_id_from_id(the_ref[0]))


def record_rejected_paper(contrib_id):
    global papers_rejected
    file = papers_rejected
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    rejected_papers = wb_obj.active

    contrib_data=find_contrib(the_id=contrib_id)
    print('db_id',contrib_data['db_id'])
    paper=get_paper_info(contrib_data['db_id'])
    
    irow=1
    if not "Contrib ID" in rejected_papers.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    irow=2
    while (rejected_papers.cell(row=irow,column=1).value is not None):
        irow=irow+1
    
    rejected_papers.cell(row=irow,column=1).value=contrib_id
    rejected_papers.cell(row=irow,column=2).value=contrib_data['db_id']
    rejected_papers.cell(row=irow,column=3).value=paper['contribution']['title']
    rejected_papers.cell(row=irow,column=4).value=paper['last_revision']['submitter']['full_name']
    rejected_papers.cell(row=irow,column=5).value=paper['last_revision']['submitter']['email']
    rejected_papers.cell(row=irow,column=6).value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb_obj.save(file)

def judge_paper(paper_db_id, decision,comment):
    print("Judging paper ", paper_db_id)
    headers = {'Authorization': f'Bearer {api_token}'}
    if decision == "To be corrected" or decision == "Minor revisions (no second round)":
        payload = {'action': 'to_be_corrected' ,  'comment': comment}
    elif decision == "accept":
        payload = {'action': 'accept' ,  'comment': comment}
    elif decision == "reject":
        payload = {'action': 'reject' ,  'comment': comment}
    else:
        print("Unable to act on this decision")
    data = requests.post(f'{event_url}papers/api/'+str(paper_db_id)+'/judge', headers=headers, data=payload)

    print(data)
    print(data.status_code)
    #if not (data.status_code == 200):
    #    exit()

    
def comment_paper(paper_db_id, comment):
    print("Commenting paper ", paper_db_id)
    headers = {'Authorization': f'Bearer {api_token}'}
    payload = { 'comment': comment}
    data = requests.post(f'{event_url}papers/api/'+str(paper_db_id)+'/comment', headers=headers, data=payload)

    print(data)
    print(data.status_code)
    #if not (data.status_code == 200):
    #    exit()

def reopen_paper(paper_db_id):
    headers = {'Authorization': f'Bearer {api_token}'}
    data = requests.delete(f'{event_url}papers/api/'+str(paper_db_id), headers=headers)
    print('Submission reopened for paper ',paper_db_id, data.status_code)
    print(data, flush=True)


def check_resubmitted_paper(the_paper):    
    global all_actions_todo
    
    secound_round_data=get_paper_in_double_reviews(the_paper['contribution']['friendly_id'])
    new_data={}
    if int(the_paper['last_revision']['number'])>int(secound_round_data['revision']):
        if secound_round_data['date_resubmitted'] is None or len(secound_round_data['date_resubmitted'])<5:

            print("Recording new version of paper ", the_paper['contribution']['friendly_id'])
            new_data['date_resubmitted']=the_paper['last_revision']['submitted_dt']
        the_ref_assigned=[]
        if secound_round_data['decision_1st_round']=="To be corrected" and (secound_round_data['suggestion_spb_2nd_round'] is None or len(secound_round_data['suggestion_spb_2nd_round'])<5 ):
            the_ref_assigned=check_referees_for_paper(the_paper['contribution']['friendly_id'])
            print('the_ref_assigned', the_ref_assigned)
            n_reviews=0        
            for timel in the_paper['last_revision']['timeline']:
                if timel['timeline_item_type']=="review":
                    n_reviews=n_reviews+1
                    the_ref_id=find_referee_by_email(timel['user']['email'])
                    print("Review by ", timel['user']['email'], the_ref_id[0])
                    if not len(the_ref_id)==1:
                        print("Error identifying referee ",timel['user']['email'], the_ref_id[0])
                        exit()
                    else:
                        refs_id=[ int(rid[0]) for rid in the_ref_assigned[1]]
                        print(refs_id)
                        if not int(the_ref_id[0]) in refs_id:
                            print("Error referee ",timel['user']['email'], the_ref_id[0]," is not assigned to this paper in the referees file ")
                        else:
                            ref_idx=refs_id.index(the_ref_id[0])+1
                            if secound_round_data['date_2nd_review'+str(ref_idx)] is not None and len(secound_round_data['date_2nd_review'+str(ref_idx)])>5 and secound_round_data['date_2nd_review'+str(ref_idx)]==timel['created_dt']:
                                #review was known
                                pass
                            else:
                                ref_dict=get_referee_dict_by_id(the_ref_id[0])
                                print(ref_dict)
                                new_data['date_2nd_review'+str(ref_idx)]=timel['created_dt']
                                new_data['name_2nd_review'+str(ref_idx)]=ref_dict['name']
                                new_data['decision_2nd_review'+str(ref_idx)]=timel['proposed_action']['name']
                                new_data['status_rev'+str(ref_idx)]="Review received"
                                new_data['date_status_rev'+str(ref_idx)]=timel['created_dt']
                                print("2nd round review received for paper ", the_paper['contribution']['friendly_id'], " from ", timel['user']['email'], the_ref_id[0],ref_dict['name']," Decision ",timel['proposed_action']['name'] )
                                mail_thank_you_review(the_paper['contribution']['friendly_id'],the_ref_id[0],the_paper['contribution']['title'],the_round=2)
                                if timel['proposed_action']['name'] not in [ 'accept' , 'reject' ]:
                                    print("Proposed action is ", timel['proposed_action']['name'] , " which is not acceptable on secound round ")
                                    all_actions_todo.append(["Proposed action is not acceptable for second round "+str(timel['proposed_action']['name']) , "./find_paper.py --id "+str(the_paper['contribution']['friendly_id'])])        
                                    #time.sleep(1)
        elif secound_round_data['decision_1st_round']=="Minor revisions (no second round)" and (secound_round_data['suggestion_spb_2nd_round'] is None or len(secound_round_data['suggestion_spb_2nd_round'])<5 ):
            if len(secound_round_data['suggestion_spb_2nd_round'])<3:
                all_actions_todo.append(["Paper with minor corrections resubmitted", "./find_paper.py --id "+str(the_paper['contribution']['friendly_id'])])
        else:
            print("Paper resubmitted decision_1st_round' was",secound_round_data['decision_1st_round'])
            print(secound_round_data)
            print("exit")
            exit()
        #notify the referees that a new revision has been submitted
        if 'date_resubmitted' in new_data and secound_round_data['decision_1st_round']=="To be corrected":
            new_data['date_notified_2nd_round']=-1
            substitution_dict={}
            print("Paper resubmitted, previous revision was ",secound_round_data['revision'])
            substitution_dict['reviewers_comments']=paper_review_report_txt(the_paper,for_spb=False,revision_number=secound_round_data['revision'])
            for idx in range(0,len(the_ref_assigned[1])):
                if 'date_2nd_review'+str(idx+1) not in new_data:
                    #print('the_ref_assigned[1][0]',the_ref_assigned[1][idx][0])
                    print("email to referees on resubmitted messages desactivated")
                    time.sleep(60)
                    #ef.send_email(referee_id=the_ref_assigned[1][idx][0],paper_id=the_paper['contribution']['friendly_id'],msgfile='message_paper_resubmitted.txt', substitution_dict=substitution_dict)
                    new_data['date_notified_2nd_round']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    new_data['status_rev'+str(idx+1)]="Notified"
                    new_data['date_status_rev'+str(idx+1)]=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        #print('new_data',new_data)
        if len(new_data.keys())>0:
            write_paper_in_double_reviews(the_paper['contribution']['friendly_id'],new_data)
            #print("resubmission recorded")
            #exit()
        
        #check two reviews
        secound_round_data=get_paper_in_double_reviews(the_paper['contribution']['friendly_id'])
        all_reviews_received=True
        if secound_round_data['decision_1st_round']=="To be corrected":
            for idx in range(0,len(the_ref_assigned[1])):
                if idx<3:
                    if not (secound_round_data['date_2nd_review'+str(idx+1)] is not None and len(secound_round_data['date_2nd_review'+str(idx+1)])>5):
                        all_reviews_received=False
        if all_reviews_received:
            all_actions_todo.append(["All reviews received" , "./find_paper.py --id "+str(the_paper['contribution']['friendly_id'])])        
            print('All actions to do so far\n',all_actions_todo)

            
def mail_thank_you_review(contrib_id,the_ref_id,paper_title,the_round=1):
                entry_code=str(((int(the_ref_id)*10000)+int(contrib_id))*7)
                urlassign="http://nicolas.delerue.org/ipac23/assign.php?"
                urlassign=urlassign+urllib.parse.urlencode({'key': str(entry_code), 'title': paper_title, 'name' : the_ref_id })                
                data = requests.get(urlassign)
                if not data.status_code == 200:
                    print("Error submitting the assignememnt: wrong response code")
                url="http://nicolas.delerue.org/ipac23/referee_additional_review.php?id="+str(entry_code)
                if the_round==2:
                    ef.send_email(referee_id=int(the_ref_id),paper_id=contrib_id,msgfile='message_thank_you_review_round2.txt',url=url)
                else:
                    ef.send_email(referee_id=int(the_ref_id),paper_id=contrib_id,msgfile='message_thank_you_review.txt',url=url)

        
def check_double_reviews_validated():
    global papers_double_review
    global days_to_notify_authors_after_1st_round
    file = papers_double_review
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    double_review_wb = wb_obj.active
    #print('Checking contrib_id in double review',contrib_id)
    
    irow=1
    if not "Contrib ID" in double_review_wb.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    irow=2
    while (double_review_wb.cell(row=irow,column=1).value is not None):
        if (double_review_wb.cell(row=irow,column=9).value is None or not len(str(double_review_wb.cell(row=irow,column=9).value))>3):
            age_notif=datetime.datetime.now()-datetime.datetime.strptime(str(double_review_wb.cell(row=irow,column=8).value),"%d/%m/%Y %H:%M")
            print('age_notif',age_notif)
            if age_notif.days >=days_to_notify_authors_after_1st_round:
                contrib_id=double_review_wb.cell(row=irow,column=1).value
                print("Contrib ",contrib_id," is ready for the second round:")
                print("./find_paper.py --id ",contrib_id," --notify-authors ")
                all_actions_todo.append(["Contrib "+str(contrib_id)+" is ready for the second round:","./find_paper.py --id "+str(contrib_id)+" --notify-authors"])
                print(all_actions_todo)
                #exit()
        irow=irow+1


def print_stats_double_reviews():
    n_papers_in_double_review_file=0
    n_papers_to_be_corrected_1st_round=0
    n_papers_minor_1st_round=0
    n_papers_accepted_1st_round=0
    n_papers_rejected_1st_round=0
    n_papers_under_spb_consideration=0
    n_papers_notified_1st_round=0
    n_papers_prab_yes=0
    
    global papers_double_review
    file = papers_double_review
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    double_review_wb = wb_obj.active
    #print('Checking contrib_id in double review',contrib_id)
    
    irow=1
    if not "Contrib ID" in double_review_wb.cell(row=irow,column=PAPER_COL).value:
        print("Error checking ", file)
        exit()
    irow=2
    while (double_review_wb.cell(row=irow,column=1).value is not None):
        n_papers_in_double_review_file=n_papers_in_double_review_file+1
        if double_review_wb.cell(row=irow,column=6).value=="To be corrected":
            n_papers_to_be_corrected_1st_round=n_papers_to_be_corrected_1st_round+1
        elif double_review_wb.cell(row=irow,column=6).value=="Minor revisions (no second round)":
            n_papers_minor_1st_round=n_papers_minor_1st_round+1
        elif double_review_wb.cell(row=irow,column=6).value=="accept":
            n_papers_accepted_1st_round=n_papers_accepted_1st_round+1
        elif double_review_wb.cell(row=irow,column=6).value=="reject":
            n_papers_rejected_1st_round=n_papers_rejected_1st_round+1
        else:
            print("Unknow paper status line",irow)
            time.sleep(10)
            exit()
        if double_review_wb.cell(row=irow,column=9).value is not None and len(double_review_wb.cell(row=irow,column=9).value) > 5:
            n_papers_notified_1st_round=n_papers_notified_1st_round+1
        else:
            n_papers_under_spb_consideration=n_papers_under_spb_consideration+1

        if double_review_wb.cell(row=irow,column=7).value=="Yes":
            n_papers_prab_yes=n_papers_prab_yes+1
        irow=irow+1
    print("* Stats from the double review file *")
    print('n_papers_in_double_review_file',n_papers_in_double_review_file)
    print('n_papers_to_be_corrected_1st_round',n_papers_to_be_corrected_1st_round)
    print('n_papers_minor_1st_round',n_papers_minor_1st_round)
    print('n_papers_accepted_1st_round',n_papers_accepted_1st_round)
    print('n_papers_rejected_1st_round',n_papers_rejected_1st_round)
    print('n_papers_under_spb_consideration',n_papers_under_spb_consideration)
    print('n_papers_notified_1st_round',n_papers_notified_1st_round)
    print('n_papers_prab_yes',n_papers_prab_yes)

def get_reviews(the_paper_revision):
    the_reviews=[]
    for timel in the_paper_revision['timeline']:
        if  timel['timeline_item_type']=="review":
            the_reviews.append(timel)
    return the_reviews

def paper_review_report_txt(the_paper,for_spb=False,revision_number=1,add_review=None,require_two_reviews=True):
    #print("paper_review_report_txt",the_paper['contribution']['id'],revision_number)
    the_reviews=get_reviews(the_paper['revisions'][int(revision_number)-1])    
    contrib=find_contrib(the_db_id=the_paper['contribution']['id'])    
    msg=""
    msg=msg+"The paper below has been reviewed by "+str(len(the_reviews))+" referees.\n\n"
    msg=msg+"Title: "+str(the_paper['contribution']['title'])+"\n"
    msg=msg+"Main Classification: "+str(contrib['track'])+"\n"
    msg=msg+"Paper code: "+str(the_paper['contribution']['code'])+"\n"
    msg=msg+"Paper url: {event_url}papers/"+str(the_paper['contribution']['id'])+"/"+"\n"
    if for_spb:
        msg=msg+'Submitter: '+str(the_paper['last_revision']['submitter']['full_name'])+' '+str(the_paper['last_revision']['submitter']['email'])+"\n"

    msg=msg+"\n***** Reviewer's report: *****\n"

    if not len(the_reviews)>=2:
        print("Paper not reviewed by two reviewers")
        if add_review is None and require_two_reviews:
            exit()    
    if len(the_reviews)>=2 or (add_review is not None and len(the_reviews)>0):
        for irat in range(len(the_reviews[0]['ratings'])):
            if "PR-AB" in the_reviews[0]['ratings'][irat]['question']['title'] or "editors"  in the_reviews[0]['ratings'][irat]['question']['title'] or "originality"  in the_reviews[0]['ratings'][irat]['question']['title'] or "can be accepted"  in the_reviews[0]['ratings'][irat]['question']['title'] or "accepted after corrections?"  in the_reviews[0]['ratings'][irat]['question']['title']:
                if for_spb:
                    msg=msg+"\n"+the_reviews[0]['ratings'][irat]['question']['title']+"\n"
                    for irev in range(0,len(the_reviews)):
                        msg=msg+"Reviewer "+str(irev+1)+": "+str(the_reviews[irev]['ratings'][irat]['value'])+"\n"
            else:
                msg=msg+"\n"+the_reviews[0]['ratings'][irat]['question']['title']+"\n"            
                for irev in range(0,len(the_reviews)):
                    msg=msg+"Reviewer "+str(irev+1)+": "+str(the_reviews[irev]['ratings'][irat]['value'])+"\n"
    if len(the_reviews)>0:
        msg=msg+"\nComments:\n\n"
    for irev in range(0,len(the_reviews)):
        if len(the_reviews[irev]['comment'])>2:
            msg=msg+"Reviewer "+str(irev+1)+": "+str(the_reviews[irev]['comment'])+"\n"

    if for_spb:
        msg=msg+"\n\nPropositions:\n"
        for irev in range(0,len(the_reviews)):
            msg=msg+the_reviews[irev]['proposed_action']['name']+"\n"

    if add_review is not None:
        msg=msg+"Additional reviews were received indirectly:\n"
        for irev in add_review:
            msg=msg+str(irev)+"\n"
        msg=msg+"\n"
    return msg


def check_contrib_consistency(contrib_db_id,the_judges,the_reviewer):
    global all_actions_todo
    global REVIEW_STATUS_COL
    global REVIEW_ACTION_COL
    global REVIEW_DATE_COL
    global REF_ID_COL
    global PAPER_COL
    global REF_NAME_COL

    #print('Consistency check')
    #print('contrib_db_id',contrib_db_id,'the_judges',the_judges,'the_reviewer',the_reviewer)
    contrib=find_contrib(the_db_id=contrib_db_id)
    contrib_id=contrib['id']
    ###Judges
    if not (len(the_judges.split("'"))==1 or len(the_judges.split("'"))==5):
        print("Incorrect number of judges")
        exit()
    #if not "nicolas delerue" in the_judges.split(",")[0].lower():
    #    print("Incorrect judge",)
    #    exit()

    ### Reviewer    
    the_ref=check_referees_for_paper(contrib_id)
    #print(the_ref)
    if len(the_reviewer)<3:
        the_reviewers_cleaned=[]
    else:
        the_reviewers_cleaned=[therev.replace("'","").strip().replace("&#39;","'") for therev in the_reviewer.split(",") ]
    #print('the_reviewers_cleaned',the_reviewers_cleaned,len(the_reviewers_cleaned))
    ref_error=False
    if the_ref[0]<len(the_reviewers_cleaned):
        print("Too many reviewers assigned")
        this_todo=[]
        this_todo.append("Too many reviewers assigned")
        for the_rev in the_reviewers_cleaned:
            #print(the_rev)
            if len(find_referee_by_name(the_rev.split(" ")[1]))>1:
                all_candidates=find_referee_by_name(the_rev.split(" ")[1])
                for cand in all_candidates:
                    #print(cand)
                    print(get_referee_by_id(cand))
                    print("Suggestion:\n", "./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(cand)+" --unassign")
                    this_todo.append([ get_referee_by_id(cand) , "./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(cand)+" --unassign"])
            else:    
                print(find_referee_by_name(the_rev.split(" ")[1]))
                print("Suggestion:\n", "./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(find_referee_by_name(the_rev.split(" ")[1])[0])+" --unassign")
                this_todo.append([ find_referee_by_name(the_rev.split(" ")[1]) , "./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(find_referee_by_name(the_rev.split(" ")[1])[0])+" --unassign" ])
        ref_error=True
        all_actions_todo.append(this_todo)
    if the_ref[0]>len(the_reviewers_cleaned):
        print("Too few reviewers assigned")
        ref_error=True
    for the_ref_data in the_ref[1]:
        if not the_ref_data[1].lower() in the_reviewers_cleaned:
            print(the_ref_data[1].lower(),' is not in ',the_reviewers_cleaned)
            print(the_ref_data[0])
            print("Suggestion: \n", "./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(the_ref_data[0])+" --assign-only")
            ref_error=True
            all_actions_todo.append("./assign_referee.py --paper "+str(contrib_id)+" --referee "+str(the_ref_data[0])+" --assign-only")
    if ref_error:
        print('contrib_id',contrib_id)
        print('Referees from excel file:',the_ref)
        print('Referees from online',the_reviewers_cleaned)
        all_actions_todo.append(["Referee error on contrib "+str(contrib_id) , 'Referees from excel file: '+str(the_ref) , 'Referees from online '+str(the_reviewers_cleaned)])
        #time.sleep(10)
    else:
        print('Consistency OK for contrib ',contrib_id)
    
def check_papers_for_referee(referee_id):
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1

    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    retval={}

    same_referee=[]
    n_papers=0
    #check entries about this paper and this referees
    irow=2
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        #print(referees_assignation_wb.cell(row=irow,column=1).value)
        if int(referees_assignation_wb.cell(row=irow,column=3).value) == int(referee_id):
            #print("Same referee")
            #print(referees_assignation_wb.cell(row=irow,column=6).value)
            if not (referees_assignation_wb.cell(row=irow,column=6).value == "Declined"):
                same_referee.append("ID: "+str(referees_assignation_wb.cell(row=irow,column=1).value)+"   Title: "+str(referees_assignation_wb.cell(row=irow,column=2).value))
                n_papers=n_papers+1
        irow=irow+1

    retval['n_papers']=n_papers
    retval['list_papers']=same_referee
    return retval

def check_history_for_referee(referee_id):
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1

    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    history=[]
    #check entries about this paper and this referees
    irow=2
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        #print(referees_assignation_wb.cell(row=irow,column=1).value)
        if int(referees_assignation_wb.cell(row=irow,column=3).value) == int(referee_id):
            the_entry={}
            the_entry['paper_id']=referees_assignation_wb.cell(row=irow,column=1).value
            the_entry['status']=referees_assignation_wb.cell(row=irow,column=6).value
            the_entry['date_status']=referees_assignation_wb.cell(row=irow,column=7).value
            the_entry['paper_title']=referees_assignation_wb.cell(row=irow,column=2).value
            
            history.append(the_entry)
        irow=irow+1

    return history  

def check_referees_for_paper(paper_id,show_row=False):
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    same_paper=[]
    n_referees=0
    #check entries about this paper and this referees
    #print('paper_id',paper_id)
    irow=2
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        #print(referees_assignation_wb.cell(row=irow,column=1).value)
        if int(referees_assignation_wb.cell(row=irow,column=1).value) == int(paper_id):
            #print("Same referee",irow)        
            if not (referees_assignation_wb.cell(row=irow,column=6).value == "Declined"):
                same_paper.append([str(referees_assignation_wb.cell(row=irow,column=3).value).strip() , str(referees_assignation_wb.cell(row=irow,column=4).value).strip()])
                n_referees=n_referees+1
                if show_row:
                    print('Paper', paper_id,  'is on row ', irow)
        irow=irow+1

    return [n_referees,same_paper]    

def check_paper_for_declined(paper_id):
    global papers_referee_assignation
    file = papers_referee_assignation
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    referees_assignation_wb = wb_obj.active

    irow=1

    paper_id=str(paper_id)
    if not "Paper id" in referees_assignation_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    list_declined=[]
    n_declined=0
    #check entries about this paper and this referees
    irow=1
    while (referees_assignation_wb.cell(row=irow,column=1).value is not None):
        if referees_assignation_wb.cell(row=irow,column=1).value == paper_id:
            if referees_assignation_wb.cell(row=irow,column=6).value == "Declined":
                print("Referee has declined", referees_assignation_wb.cell(row=irow,column=3).value, referees_assignation_wb.cell(row=irow,column=4).value)
                list_declined.append(referees_assignation_wb.cell(row=irow,column=3).value)
                n_declined=n_declined+1
        irow=irow+1

    return [n_declined,list_declined]
    
######

#### Registration database
def find_participant_by_name(name):
    file = "registrations_20230323.xlsx"
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    registration_wb = wb_obj.active

    irow=1

    if not "ID" in registration_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    #check entries about this paper and this referees
    irow=2
    while (registration_wb.cell(row=irow,column=1).value is not None):
        if str(registration_wb.cell(row=irow,column=2).value) == str(name):            
            the_entry={}
            the_entry['id']=registration_wb.cell(row=irow,column=1).value
            the_entry['name']=registration_wb.cell(row=irow,column=2).value
            the_entry['email']=registration_wb.cell(row=irow,column=3).value
            the_entry['country']=registration_wb.cell(row=irow,column=5).value
            the_entry['category']=registration_wb.cell(row=irow,column=6).value
            the_entry['affiliation']=registration_wb.cell(row=irow,column=4).value
            return the_entry
        irow=irow+1
    return None

def find_participant_by_email(email):
    file = "registrations_20230323.xlsx"
    wb_obj = openpyxl.load_workbook(file) 

    # Read the active sheet:
    registration_wb = wb_obj.active

    irow=1

    if not "ID" in registration_wb.cell(row=irow,column=1).value:
        print("Error checking ", file)
        exit()

    #check entries about this paper and this referees
    irow=2
    while (registration_wb.cell(row=irow,column=1).value is not None):
        if str(registration_wb.cell(row=irow,column=3).value) == str(email):
            the_entry={}
            the_entry['id']=registration_wb.cell(row=irow,column=1).value
            the_entry['name']=registration_wb.cell(row=irow,column=2).value
            the_entry['email']=registration_wb.cell(row=irow,column=3).value
            the_entry['country']=registration_wb.cell(row=irow,column=5).value
            the_entry['category']=registration_wb.cell(row=irow,column=6).value
            the_entry['affiliation']=registration_wb.cell(row=irow,column=4).value
            return the_entry
        irow=irow+1
    return None

def search_user(email=None,last_name=None,first_name=None): 
        headers = {'Authorization': f'Bearer {api_token}'}
        if email is not None:
            data = requests.get(f'https://indico.jacow.org/user/search/?email='+email+'&favorites_first=true', headers=headers)
        elif last_name is not None:
            if first_name is not None:
                data = requests.get(f'https://indico.jacow.org/user/search/?last_name='+last_name+'&first_name='+first_name, headers=headers)
            else:
                data = requests.get(f'https://indico.jacow.org/user/search/?last_name='+last_name, headers=headers)

        if not data.status_code == 200:
            print("Error searching for user")
            return None
        data_json=data.json()
        if len(data_json['users'])>1:
            print("More than one user found")
            return None
        elif len(data_json['users'])==0:
            print("No user found")
            return None
        else:
            return data_json['users'][0]

def update_reviewer_map(email,ref_id):
    global allrefdata_fname
    [the_sub_MC_list,all_referees_by_sub_MC,all_referees_by_sub_MC_coauthors]=joblib.load(allrefdata_fname)
    for idx in range(0,len(all_referees_by_sub_MC)):
        if email in all_referees_by_sub_MC[idx]:
            jdx=all_referees_by_sub_MC[idx].index(email)
            all_referees_by_sub_MC[idx][jdx]=ref_id
    for idx in range(0,len(all_referees_by_sub_MC_coauthors)):
        if email in all_referees_by_sub_MC_coauthors[idx]:
            jdx=all_referees_by_sub_MC_coauthors[idx].index(email)
            all_referees_by_sub_MC_coauthors[idx][jdx]=ref_id
    joblib.dump([the_sub_MC_list,all_referees_by_sub_MC,all_referees_by_sub_MC_coauthors],allrefdata_fname)



##### Paper status ####

papers_status_file="papers_status.xlsx"

papers_status_cols_structure=[
    [ 'id' , 1]  ,
    [ 'db_id' , 2] ,
    [ 'session_code' , 3] ,
    [ 'title' , 4] ,
    [ 'track' , 5] ,
    [ 'url' , 6] ,
    [ 'submitter_name' , 10] ,
    [ 'submitter_email' , 11] ,
    [ 'submitter_affiliation' , 12] ,
    [ 'authors_countries' , 13] ,
    [ 'date_submitted' , 15] ,
    [ 'paper_status' , 16] ,
    [ 'date_status' , 17] ,
    [ 'indico_state' , 18] ,
    [ 'indico_state_date' , 19] ,
    [ 'n_declined' , 20] ,
    [ 'remarks' , 25]
]

def check_paper_in_status_file(contrib_id):
    return get_paper_in_status_file(contrib_id) is not None

def get_paper_in_status_file(contrib_id):
    global papers_status_file
    global papers_status_cols_structure
    file = papers_status_file
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    paper_status_wb = wb_obj.active
    #print('Checking contrib_id in double review',contrib_id)
    retval=None
    
    irow=1
    if not "Contrib ID" in paper_status_wb.cell(1,1).value:
        print("Error checking ", file)
        exit()
    while (paper_status_wb.cell(row=irow,column=1).value is not None) and not str(paper_status_wb.cell(row=irow,column=1).value) == str(contrib_id):
        irow=irow+1
    if str(paper_status_wb.cell(row=irow,column=1).value) == str(contrib_id):
        print("Paper found in file",file)
        retval={}
        retval['row']=irow
        for the_key in paper_status_cols_structure:
            retval[the_key[0]]=str(paper_status_wb.cell(row=irow,column=int(the_key[1])).value)
        return retval
    else:
        print("Paper not found in file",file)
        return None
    return None

def write_paper_in_papers_status(contrib_id,paper_data,add_entry=False):
    global papers_status_file
    global papers_status_cols_structure
    file = papers_status_file
    wb_obj = openpyxl.load_workbook(file) 
    # Read the active sheet:
    paper_status_wb = wb_obj.active
    #print('Checking contrib_id in double review',contrib_id)
    retval=None
    
    irow=1
    if not "Contrib ID" in paper_status_wb.cell(1,1).value:
        print("Error checking ", file)
        exit()
    while (paper_status_wb.cell(row=irow,column=1).value is not None) and not str(paper_status_wb.cell(row=irow,column=1).value) == str(contrib_id):
        irow=irow+1
    if str(paper_status_wb.cell(row=irow,column=1).value) == str(contrib_id) or add_entry:
        #print("Paper found in file - writing", file, irow)
        #print(paper_data)
        for the_key in papers_status_cols_structure:
            if the_key[0] in  paper_data.keys():
                paper_status_wb.cell(row=irow,column=int(the_key[1])).value=paper_data[the_key[0]]
            
        wb_obj.save(file)
        #print("saved")
        return irow
    else:
        print("Error: Paper not found in file",file)
        exit()
    print("Error: Paper not found in file",file)
    exit()

def check_all_papers_status():
    for the_id in range(500,2800):
        contrib=find_contrib(the_id)
        if contrib is not None:
            print('Checking paper for contrib ', the_id,' that is not none; db_id ',contrib['db_id'])
            paper_data={}
            paper_data[ 'id' ]=the_id
            paper_data[ 'db_id' ]=contrib['db_id']
            the_paper=get_paper_info(contrib['db_id'],use_cache=True,sleep_before_online=1)
            paper_data[ 'session_code' ]=contrib['code']
            paper_data[ 'title' ]=contrib['title']
            paper_data[ 'track']=contrib['track']
            paper_data[ 'url']="{event_url}papers/"+str(contrib['db_id'])+"/"
            #continent=geo_dict['continent']
            if the_paper is None:
                print("None")
                paper_data[ 'paper_status' ]="No longer available"
            else:
                geo_dict=get_countries_from_contrib(contrib)
                paper_data[ 'authors_countries' ]=str(geo_dict['countries'])
                paper_data[ 'submitter_name' ]=str(the_paper['revisions'][0]['submitter']['full_name'])
                paper_data[ 'submitter_email']=str(the_paper['revisions'][0]['submitter']['email'])
                paper_data[ 'submitter_affiliation' ]=str(the_paper['revisions'][0]['submitter']['affiliation'])
                paper_data[ 'indico_state'] = the_paper['state']['name']
                if the_paper['state']['name'] == "submitted":
                    paper_data[ 'indico_state_date'] = the_paper['last_revision']["submitted_dt"]
                    secound_round_data=get_paper_in_double_reviews(the_paper['contribution']['friendly_id'])
                    if secound_round_data is None:
                        paper_data[ 'paper_status' ]="First round of review"
                        paper_data[ 'date_status']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        if secound_round_data['date_resubmitted'] is not None and len(secound_round_data['date_resubmitted'])>5:
                            if secound_round_data['suggestion_spb_2nd_round'] is not None  and len(secound_round_data['suggestion_spb_2nd_round'])>5:
                                paper_data[ 'paper_status' ]="Final decision due soon"
                                paper_data[ 'date_status']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                paper_data[ 'paper_status' ]="Resubmitted for second round"
                                paper_data[ 'date_status']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        else:
                                paper_data[ 'paper_status' ]="Not yet resubmitted for 2nd round"
                                paper_data[ 'date_status']=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                else:
                    paper_data[ 'paper_status' ] = the_paper['state']['name']
                    paper_data[ 'date_status'] = the_paper['last_revision']["judgment_dt"]
                    paper_data[ 'indico_state_date'] = the_paper['last_revision']["judgment_dt"]
                print(paper_data)
                write_paper_in_papers_status(the_id,paper_data,add_entry=True)

